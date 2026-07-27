"""
excel_generator.py — LBO Model Excel Export (V5, Tâche "Élever l'export Excel
du LBO au standard professionnel").

Génère un classeur .xlsx à 8 onglets où TOUTE valeur dérivée est une formule
Excel native et TOUTE hypothèse est une cellule d'entrée identifiable (police
bleue) :
  1. **Cover & Assumptions** — en-tête deal (cible/secteur/pays/date) +
     provenance réel/estimé des chiffres clés, puis toutes les hypothèses.
  2. **Sources & Uses**      — emplois (EV) vs sources (dette + equity plug),
     références pures vers l'onglet 1 (aucune hypothèse dupliquée), check dur.
  3. **Operating Model**     — Revenue → EBITDA → (intérêts) → EBT → impôt →
     FCF. Le moteur ne modélise pas de D&A/EBIT distinct de l'EBITDA — la
     base imposable réelle est EBITDA − intérêts (jamais inventé un poste
     D&A que le moteur n'a pas).
  4. **Debt Schedule**       — par tranche (V3 multi-tranche) OU tranche
     unique à cash-sweep intégral (V2 legacy — c'est le mode réellement
     utilisé par les 3 deals canoniques du produit) : solde d'ouverture,
     intérêts, remboursement (formule différente selon le mode réel),
     solde de clôture.
  5. **Returns & Waterfall** — Exit EV, Exit Equity, MoIC, IRR natif (=IRR())
     et le waterfall Fund/Management si le moteur en a un.
  6. **Sensitivity**         — grilles IRR/MOIC (entrée×sortie, levier×sortie)
     calculées en RÉUTILISANT `ic_context.compute_*_sensitivity` (qui
     rappelle `run_lbo_model` — jamais une nouvelle formule). Ce sont des
     valeurs figées au moment de l'export, explicitement étiquetées comme
     telles (reconstruire une simulation multi-année complète dans 25+
     cellules Excel sans macro n'est pas praticable — limite honnête,
     documentée, pas contournée en fabriquant une fausse formule "vivante").
  7. **Credit Metrics**      — Dette/EBITDA, EBITDA/Intérêts, FCF/Dette, DSCR
     — combinaisons de cellules déjà présentes sur Operating Model/Debt
     Schedule, aucune nouvelle donnée moteur.
  8. **Checks**              — Sources=Emplois, cohérence du roll-forward de
     dette, cohérence des returns, cohérence de l'EBITDA entre onglets, pas
     de solde de dette négatif — chaque ligne PASS/FAIL par formule, jamais
     une valeur écrite en dur.

Réplique exactement l'algorithme du moteur Python (`valuation_engine.py`,
NON MODIFIÉ — seules ses constantes en lecture seule sont importées) :
  - le mode multi-tranche (V3) N'A PAS de cash-sweep au-delà de
    l'amortissement programmé (cf. `_build_lbo` : le bloc de sweep ne
    s'exécute que si `not use_multi_tranche`) — une tranche "bullet" ne
    rembourse donc JAMAIS de principal avant la sortie dans ce moteur ;
  - le mode V2 legacy (pas de `debt_structure`, seul mode réellement produit
    par `build_base_case_scenario` pour tous les deals promus aujourd'hui) a,
    à l'inverse, un cash-sweep intégral chaque année (`sweep = min(cash,
    balance)`), sans annuité fixe programmée ;
  - le remboursement programmé d'une tranche amortissable (V3) est plafonné
    par trois éléments : l'annuité fixe (montant final / durée), le solde
    restant, et le cash disponible cascadé par ordre de séniorité — les
    trois sont recalculés en Excel via MIN(), pas copiés depuis le moteur ;
  - le montant de chaque tranche est plafonné à `MAX_LEVERAGE_PCT` (60 %) de
    l'EV d'entrée — le facteur d'échelle est lui-même une formule.

Ce qui n'est PAS reconstruit en formule (limite honnête, documentée) : le
déclenchement du ratchet de management (le seuil de TRI et le bonus ne sont
pas exposés dans le payload du moteur — `management_total_pct` est repris
comme hypothèse d'entrée, pas recalculé depuis un seuil vivant) ; la grille
de sensibilité (voir onglet 6, ci-dessus) ; un « Net Debt » — le moteur ne
modélise aucun solde de trésorerie séparé, seul un `Total Debt` (brut) est
donc affiché, jamais un chiffre net inventé.

Requires: openpyxl >= 3.1
"""

from __future__ import annotations

import io
from dataclasses import dataclass, field
from datetime import datetime

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from api.services.ma_engine.valuation_engine import (
    DSCR_COVENANT_MIN,
    FINANCING_FEE_PCT,
    LEVERAGE_COVENANT_STEPDOWN_PCT,
    MAX_LEVERAGE_PCT,
    MIN_CASH_PCT,
    TRANSACTION_FEE_PCT,
)
from api.services.ma_engine.ic_context import (
    compute_entry_exit_sensitivity,
    compute_exit_leverage_sensitivity,
)


# ============================================================
# Style constants — PE / Investment Banking palette
# ============================================================

_NAVY = "1B2A4A"
_DARK_BLUE = "2C3E6B"
_LIGHT_BLUE = "D6E4F0"
_LIGHT_GREY = "F2F2F2"
_WHITE = "FFFFFF"
_GREEN = "27AE60"
_GOLD = "F39C12"
_RED = "C0392B"
_INPUT_BLUE = "0000FF"   # Convention IB classique : entrées codées en bleu
_LINK_GREEN = "007A33"   # Convention IB : liens inter-feuilles en vert

_FONT_HEADER = Font(name="Calibri", bold=True, color=_WHITE, size=11)
_FONT_SECTION = Font(name="Calibri", bold=True, color=_NAVY, size=11)
_FONT_NORMAL = Font(name="Calibri", size=10)
_FONT_BOLD = Font(name="Calibri", bold=True, size=10)
_FONT_TITLE = Font(name="Calibri", bold=True, color=_WHITE, size=14)
_FONT_KPI = Font(name="Calibri", bold=True, color=_NAVY, size=16)
_FONT_INPUT = Font(name="Calibri", size=10, color=_INPUT_BLUE, bold=True)
_FONT_INPUT_KPI = Font(name="Calibri", size=16, bold=True, color=_INPUT_BLUE)
_FONT_LINK = Font(name="Calibri", size=10, color=_LINK_GREEN, bold=True)
_FONT_LINK_BOLD = Font(name="Calibri", size=10, color=_LINK_GREEN, bold=True)
_FONT_NOTE = Font(name="Calibri", size=8, italic=True, color="777777")
_FONT_PASS = Font(name="Calibri", bold=True, size=10, color=_GREEN)
_FONT_FAIL = Font(name="Calibri", bold=True, size=10, color=_RED)

_FILL_HEADER = PatternFill(start_color=_NAVY, end_color=_NAVY, fill_type="solid")
_FILL_SECTION = PatternFill(start_color=_DARK_BLUE, end_color=_DARK_BLUE, fill_type="solid")
_FILL_LIGHT = PatternFill(start_color=_LIGHT_BLUE, end_color=_LIGHT_BLUE, fill_type="solid")
_FILL_ALT = PatternFill(start_color=_LIGHT_GREY, end_color=_LIGHT_GREY, fill_type="solid")

_THIN_BORDER = Border(bottom=Side(style="thin", color="CCCCCC"))
_THICK_BOTTOM = Border(bottom=Side(style="medium", color=_NAVY))

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
_ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)
_ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
_ALIGN_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)

_FMT_EUR = '#,##0'
_FMT_PCT = '0.00%'
_FMT_MULT = '0.00"x"'
_FMT_NUM_2 = '#,##0.00'

_SHEET_ASSUMPTIONS = "Cover & Assumptions"
_SHEET_SOURCES_USES = "Sources & Uses"
_SHEET_PNL = "Operating Model"
_SHEET_DEBT = "Debt Schedule"
_SHEET_RETURNS = "Returns & Waterfall"
_SHEET_SENSITIVITY = "Sensitivity"
_SHEET_CREDIT = "Credit Metrics"
_SHEET_CHECKS = "Checks"


# ============================================================
# Helpers
# ============================================================

def _col(n: int) -> str:
    """1-indexed column number to letter (1=A, 2=B, …)."""
    return get_column_letter(n)


def _xref(sheet: str, col: int, row: int) -> str:
    """Cross-sheet absolute reference, e.g. ='Operating Model'!$C$12."""
    return f"'{sheet}'!${_col(col)}${row}"


def _style_header_row(ws, row: int, max_col: int):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = _ALIGN_CENTER
        cell.border = _THICK_BOTTOM


def _style_section_label(ws, row: int, col: int, label: str):
    cell = ws.cell(row=row, column=col, value=label)
    cell.font = _FONT_SECTION
    cell.alignment = _ALIGN_LEFT


def _write_title(ws, text: str, n_cols: int):
    # Title banners carry long descriptive text (professional-model
    # convention: the banner doubles as a one-line "what am I looking at"
    # note) — a merge sized to the sheet's DATA columns is often narrower
    # than the text itself. Left-aligned + a generously wide merge avoids
    # the title being clipped on BOTH edges when printed/exported to PDF
    # (a centred title in a too-narrow merge overflows symmetrically and
    # gets cut on both sides at the page boundary — observed and fixed
    # during verification).
    width = max(n_cols, 14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    title_cell = ws.cell(row=1, column=1, value=text)
    title_cell.font = _FONT_TITLE
    title_cell.fill = _FILL_HEADER
    title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for c in range(1, width + 1):
        ws.cell(row=1, column=c).fill = _FILL_HEADER
    ws.freeze_panes = "B4"


def _write_input(ws, row: int, col: int, label: str, value, fmt: str | None = None,
                  comment: str | None = None, kpi: bool = False):
    """Write a labelled BLUE input cell (col A = label, given col = value)."""
    ws.cell(row=row, column=1, value=label).font = _FONT_NORMAL
    ws.cell(row=row, column=1).alignment = _ALIGN_LEFT
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = _FONT_INPUT_KPI if kpi else _FONT_INPUT
    cell.alignment = _ALIGN_RIGHT
    if fmt:
        cell.number_format = fmt
    if comment:
        cell.comment = Comment(comment, "PE Tracker")
    return cell


def _write_formula(ws, row: int, col: int, label: str, formula: str, fmt: str | None = None,
                    bold: bool = True, fill_light: bool = False, kpi: bool = False, link: bool = False,
                    comment: str | None = None):
    """Write a labelled FORMULA cell. Black/navy = ordinary computed cell;
    green (`link=True`) = a value that is PURELY a reference to another sheet
    (IB convention: blue input / black formula / green inter-sheet link)."""
    ws.cell(row=row, column=1, value=label).font = _FONT_NORMAL
    ws.cell(row=row, column=1).alignment = _ALIGN_LEFT
    cell = ws.cell(row=row, column=col, value=formula)
    if link:
        cell.font = _FONT_LINK_BOLD if bold else _FONT_LINK
    else:
        cell.font = _FONT_KPI if kpi else (_FONT_BOLD if bold else _FONT_NORMAL)
    cell.alignment = _ALIGN_RIGHT
    if fmt:
        cell.number_format = fmt
    if fill_light:
        cell.fill = _FILL_LIGHT
        ws.cell(row=row, column=1).fill = _FILL_LIGHT
    if comment:
        cell.comment = Comment(comment, "PE Tracker")
    return cell


def _auto_width(ws, min_width: float = 12, max_width: float = 30):
    # Landscape + fit-to-width: every sheet in this workbook is wider than one
    # portrait page (year columns, tranche tables) — without this, printing or
    # exporting to PDF splits a single row's title banner across two pages.
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    for col_cells in ws.columns:
        max_len = min_width
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, min(len(str(cell.value)) + 3, max_width))
        ws.column_dimensions[col_letter].width = max_len


def _provenance_summary(entry: dict | None) -> str:
    """One-line, human-readable summary of a FieldProvenance dict — real
    (REGISTRY/DOCUMENT/MARKET) vs estimated (ESTIMATE) vs manual, exactly as
    already tracked by the rest of the app (D18) — never re-derived here."""
    if not entry:
        return "origin not tracked"
    label = entry.get("provenance", "UNKNOWN")
    ref = entry.get("reference", "") or ""
    as_of = entry.get("as_of")
    prefix = {
        "REGISTRY": "REAL — official registry",
        "DOCUMENT": "REAL — uploaded document",
        "MARKET": "REAL — market data provider",
        "MANUAL": "MANUAL — analyst entry",
        "ESTIMATE": "ESTIMATED — modelled",
    }.get(label, label)
    as_of_txt = f" (FY{as_of})" if as_of else ""
    ref_txt = f" — {ref}" if ref else ""
    return f"{prefix}{as_of_txt}{ref_txt}"


# ============================================================
# Layout plan — row numbers computed BEFORE any cell is written,
# so every sheet can cross-reference every other sheet regardless
# of build order (see module docstring).
# ============================================================

@dataclass
class AssumptionsLayout:
    row_entry_revenue: int = 0
    row_ebitda_margin: int = 0
    row_revenue_growth: int = 0
    row_capex_pct: int = 0
    row_wcr_pct: int = 0
    row_tax_rate: int = 0
    row_holding_period: int = 0
    row_max_leverage: int = 0
    # Tâche "P1 : physique financière du modèle LBO" (Parties D/E) — nouvelles
    # hypothèses ESTIMATE : frais réels de Sources & Uses + seuils de covenant.
    row_transaction_fee_pct: int = 0
    row_financing_fee_pct: int = 0
    row_min_cash_pct: int = 0
    row_dscr_covenant_min: int = 0
    row_leverage_covenant_stepdown: int = 0
    has_calibration: bool = False
    row_calib_median: int = 0
    row_calib_discount: int = 0
    row_calib_derived: int = 0
    row_entry_multiple: int = 0
    row_exit_multiple: int = 0
    row_entry_ebitda: int = 0
    row_ev: int = 0
    row_leverage_scale: int = 0
    row_leverage_turns_v2: int = 0  # V2 legacy only — turns input; Total Debt derives from it
    row_total_debt: int = 0
    row_interest_rate_v2: int = 0  # V2 legacy only — no per-tranche table
    row_transaction_fees: int = 0
    row_financing_fees: int = 0
    row_min_cash: int = 0
    row_uses_total: int = 0
    row_sponsor_equity: int = 0
    row_leverage_mult: int = 0
    n_tranches: int = 0
    tranche_rows: list[int] = field(default_factory=list)  # one row per tranche in the input table
    # Debt table columns (fixed): A=name B=turns C=rate D=amort-type E=raw amount F=final amount
    COL_TURNS = 2
    COL_RATE = 3
    COL_AMORT_TYPE = 4
    COL_RAW_AMOUNT = 5
    COL_FINAL_AMOUNT = 6


def _plan_assumptions(res: dict, start_row: int = 4) -> AssumptionsLayout:
    lay = AssumptionsLayout()
    calibration = res.get("calibration")
    lay.has_calibration = bool(
        calibration and calibration.get("sufficient") and calibration.get("applicable")
    )
    tranches = res.get("debt_tranches_detail") or []
    lay.n_tranches = len(tranches)

    r = start_row
    lay.row_entry_revenue = r; r += 1
    lay.row_ebitda_margin = r; r += 1
    lay.row_revenue_growth = r; r += 1
    lay.row_capex_pct = r; r += 1
    lay.row_wcr_pct = r; r += 1
    lay.row_tax_rate = r; r += 1
    lay.row_holding_period = r; r += 1
    lay.row_max_leverage = r; r += 1
    lay.row_transaction_fee_pct = r; r += 1
    lay.row_financing_fee_pct = r; r += 1
    lay.row_min_cash_pct = r; r += 1
    lay.row_dscr_covenant_min = r; r += 1
    lay.row_leverage_covenant_stepdown = r; r += 1

    r += 2  # section header
    if lay.has_calibration:
        lay.row_calib_median = r; r += 1
        lay.row_calib_discount = r; r += 1
        lay.row_calib_derived = r; r += 1
        r += 1
    lay.row_entry_multiple = r; r += 1
    lay.row_exit_multiple = r; r += 1

    r += 2  # section header
    lay.row_entry_ebitda = r; r += 1
    lay.row_ev = r; r += 1
    if lay.n_tranches:
        lay.row_leverage_scale = r; r += 1
    else:
        lay.row_leverage_turns_v2 = r; r += 1
    lay.row_total_debt = r; r += 1
    if not lay.n_tranches:
        lay.row_interest_rate_v2 = r; r += 1
    lay.row_transaction_fees = r; r += 1
    lay.row_financing_fees = r; r += 1
    lay.row_min_cash = r; r += 1
    lay.row_uses_total = r; r += 1
    lay.row_sponsor_equity = r; r += 1
    lay.row_leverage_mult = r; r += 1

    if lay.n_tranches:
        r += 2  # section header + column header
        for _ in range(lay.n_tranches):
            lay.tranche_rows.append(r)
            r += 1

    return lay


@dataclass
class DebtScheduleLayout:
    # Tâche "P1 : physique financière du modèle LBO" (Parties B/D) — 9 lignes
    # par bloc de tranche (contre 5 avant) : le remboursement programmé
    # (Erreur 2) et le cash sweep (Erreur 1) sont désormais DEUX passes
    # explicites et distinctes, chacune avec son propre "cash disponible"
    # cascadé par séniorité — au lieu d'une seule ligne "repayment" qui
    # mélangeait les deux (et qui, en V3, ignorait purement et simplement le
    # cash sweep sur les tranches bullet).
    tranche_blocks: list[dict[str, int]] = field(default_factory=list)
    row_total_debt_eoy: int = 0
    row_total_interest: int = 0
    row_total_sched_due: int = 0
    row_total_sched_paid: int = 0
    row_total_shortfall: int = 0
    row_total_repayment: int = 0
    row_cash_balance_eoy: int = 0
    row_net_debt_eoy: int = 0


def _plan_debt_schedule(n_blocks: int) -> DebtScheduleLayout:
    lay = DebtScheduleLayout()
    r = 4  # row 3 = header
    for _ in range(n_blocks):
        block = {
            "cash_avail_sched": r,
            "opening": r + 1,
            "interest": r + 2,
            "sched_due": r + 3,
            "sched_paid": r + 4,
            "cash_avail_sweep": r + 5,
            "sweep_repay": r + 6,
            "total_repay": r + 7,
            "closing": r + 8,
        }
        lay.tranche_blocks.append(block)
        r += 10  # 9 rows + 1 blank
    r += 1
    lay.row_total_debt_eoy = r; r += 1
    lay.row_total_interest = r; r += 1
    lay.row_total_sched_due = r; r += 1
    lay.row_total_sched_paid = r; r += 1
    lay.row_total_shortfall = r; r += 1
    lay.row_total_repayment = r; r += 1
    lay.row_cash_balance_eoy = r; r += 1
    lay.row_net_debt_eoy = r
    return lay


@dataclass
class PnlLayout:
    row_revenue: int = 0
    row_ebitda: int = 0
    row_da: int = 0
    row_ebit: int = 0
    row_interest: int = 0
    row_taxable_income: int = 0
    row_tax: int = 0
    row_capex: int = 0
    row_wcr: int = 0
    row_fcf: int = 0
    row_debt_repayment: int = 0
    row_debt_eoy: int = 0
    row_cash_eoy: int = 0
    row_net_debt_eoy: int = 0


def _plan_pnl() -> PnlLayout:
    lay = PnlLayout()
    r = 5  # row 3 = title/header, row 4 = "P&L" section label
    lay.row_revenue = r; r += 1
    lay.row_ebitda = r; r += 1
    lay.row_da = r; r += 1
    lay.row_ebit = r; r += 1
    lay.row_interest = r; r += 1
    lay.row_taxable_income = r; r += 1
    lay.row_tax = r; r += 1
    r += 1  # "CASH FLOW" section label
    lay.row_capex = r; r += 1
    lay.row_wcr = r; r += 1
    lay.row_fcf = r; r += 1
    r += 1  # "DEBT SCHEDULE" section label
    lay.row_debt_repayment = r; r += 1
    lay.row_debt_eoy = r; r += 1
    lay.row_cash_eoy = r; r += 1
    lay.row_net_debt_eoy = r
    return lay


# ============================================================
# Main generator
# ============================================================

def generate_lbo_model_excel(
    lbo_result: dict, *, deal=None, scenario_label: str | None = None,
    sizing_note: str | None = None,
    downside_result: dict | None = None, downside_label: str | None = None,
) -> io.BytesIO:
    """Generate a fully-formulaic, 8-tab professional LBO model workbook from
    engine output.

    Args:
        lbo_result: dict returned by run_lbo_model() (V3-compatible), with
            optional "calibration" / "financial_provenance" keys (D22/D45)
            for provenance display.
        deal: optional `Deal` ORM instance — when given, the Cover block
            (target/acquirer/sector/country/scenario/date) is written at the
            top of the Cover & Assumptions tab. Omitted (None) for the
            deal-agnostic manual calculator export, which keeps its exact
            pre-existing layout.
        scenario_label: optional label of the LBOScenario this export was
            generated from (e.g. "Base case (auto)") — cosmetic only.
        sizing_note: Tâche "P2" (Partie A) — when the target is under the
            small-cap threshold, the same indicative/bolt-on caveat shown in
            the memo/deck, rendered on the Cover tab.
        downside_result: Tâche "P2" (Partie B) — the deal's downside
            scenario's `result_json`, if one exists, shown as a compact
            snapshot on the Cover tab (never a second live model).
        downside_label: label of the downside LBOScenario (cosmetic).

    Returns:
        io.BytesIO containing the .xlsx bytes, ready for streaming.
    """
    res = lbo_result
    projs = res.get("projections", [])
    n_years = len(projs)  # includes Year 0
    tranches = res.get("debt_tranches_detail") or []
    has_tranches = bool(tranches)
    is_v2_debt = (not has_tranches) and (res.get("entry_debt") or 0) > 0
    has_debt_schedule = (has_tranches or is_v2_debt) and bool(projs)

    wb = Workbook()

    cover_header_row = _write_cover_block(
        wb, deal, scenario_label, res,
        sizing_note=sizing_note, downside_result=downside_result, downside_label=downside_label,
    )
    assum_layout = _plan_assumptions(res, start_row=cover_header_row + 1)
    debt_layout = (
        _plan_debt_schedule(assum_layout.n_tranches if has_tranches else 1)
        if has_debt_schedule else None
    )
    pnl_layout = _plan_pnl() if projs else None

    _build_assumptions_sheet(wb, res, assum_layout, cover_header_row)
    su_layout = _build_sources_uses_sheet(wb, res, assum_layout)
    if projs:
        _build_projections_sheet(wb, res, pnl_layout, assum_layout, debt_layout, n_years)
    if has_debt_schedule:
        _build_debt_schedule_sheet(wb, res, debt_layout, assum_layout, pnl_layout, n_years, is_v2=is_v2_debt)
    returns_layout = _build_returns_sheet(wb, res, assum_layout, pnl_layout, debt_layout, n_years)

    entry_exit_sens = compute_entry_exit_sensitivity(res.get("entry_revenue"), "", res)
    exit_lev_sens = compute_exit_leverage_sensitivity(res.get("entry_revenue"), "", res)
    if entry_exit_sens or exit_lev_sens:
        _build_sensitivity_sheet(wb, entry_exit_sens, exit_lev_sens)

    credit_layout = None
    if has_debt_schedule and pnl_layout is not None:
        credit_layout = _build_credit_metrics_sheet(wb, res, assum_layout, pnl_layout, debt_layout, n_years)

    _build_checks_sheet(
        wb, res, assum_layout, pnl_layout, debt_layout, returns_layout, credit_layout,
        n_years, has_debt_schedule, is_v2_debt,
    )

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ============================================================
# Sheet 1 — Cover & Assumptions (ALL inputs, blue; every derived cell = formula)
# ============================================================

def _write_cover_block(
    wb: Workbook, deal, scenario_label: str | None, res: dict, *,
    sizing_note: str | None = None,
    downside_result: dict | None = None, downside_label: str | None = None,
) -> int:
    """Creates the Cover & Assumptions sheet and writes the deal header +
    data-provenance block (rows 2..N), if `deal` is given. Returns the row
    where the "OPERATING ASSUMPTIONS" section header belongs — 3 (the
    pre-existing, unchanged layout) when `deal` is None, so the deal-agnostic
    manual-calculator export is byte-for-byte unaffected by this tab's
    rename/extension.

    Tâche "P2 : crédibilité de la thèse" — `sizing_note` (Partie A) and
    `downside_result`/`downside_label` (Partie B) are optional, additive
    blocks: when the target is under the small-cap threshold, a caveat is
    shown here (same wording as the memo/deck) ; when a downside scenario
    exists for this deal, a compact SNAPSHOT of its key figures is shown —
    a static summary (like the Sensitivity tab), not a second live model."""
    ws = wb.create_sheet(_SHEET_ASSUMPTIONS, 0)
    ws.sheet_properties.tabColor = _NAVY
    title = f"LBO MODEL — {deal.target_name.upper()}" if deal and deal.target_name else "LBO MODEL — ASSUMPTIONS"
    _write_title(ws, f"{title} (blue = input, black = formula, green = link)", 6)

    if deal is None:
        return 3

    r = 2
    cell = ws.cell(row=r, column=1, value=(
        f"Target: {deal.target_name or 'N/A'}   |   Acquirer: {deal.acquirer_name or 'N/A'}   |   "
        f"Sector: {deal.sector or 'N/A'}   |   Country: {deal.country or 'N/A'}"
    ))
    cell.font = _FONT_BOLD
    r += 1
    cell = ws.cell(row=r, column=1, value=(
        f"Scenario: {scenario_label or 'N/A'}   |   Generated: {datetime.now().strftime('%d %B %Y')}"
    ))
    cell.font = _FONT_NOTE
    r += 2

    if sizing_note:
        _style_section_label(ws, r, 1, "SIZING CAVEAT — INDICATIVE MODEL")
        r += 1
        note_cell = ws.cell(row=r, column=1, value=sizing_note)
        note_cell.font = Font(name="Calibri", size=9, italic=True, color=_RED)
        note_cell.alignment = _ALIGN_WRAP
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.row_dimensions[r].height = 60
        r += 2

    prov = res.get("financial_provenance") or {}
    _style_section_label(ws, r, 1, "DATA PROVENANCE — real vs estimated")
    r += 1
    for key, label in (
        ("revenue", "Entry Revenue"),
        ("ebitda_margin", "EBITDA Margin / Entry EBITDA"),
        ("entry_multiple", "Entry Multiple"),
    ):
        cell = ws.cell(row=r, column=1, value=f"  {label}: {_provenance_summary(prov.get(key))}")
        cell.font = _FONT_NOTE
        r += 1
    r += 1

    if downside_result:
        _style_section_label(ws, r, 1, f"{(downside_label or 'DOWNSIDE CASE').upper()} — SNAPSHOT (not a live model)")
        r += 1
        note_cell = ws.cell(row=r, column=1, value=(
            "Computed by the same engine (run_lbo_model) with the base case's own frozen "
            "sector profile — only entry revenue and exit multiple are degraded. A static "
            "snapshot at export time, like the Sensitivity tab; see the Returns Analysis "
            "section of the memo/deck for the base-vs-downside comparison."
        ))
        note_cell.font = _FONT_NOTE
        note_cell.alignment = _ALIGN_WRAP
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.row_dimensions[r].height = 30
        r += 1
        downside_fields = [
            ("Entry Revenue", _FMT_EUR, downside_result.get("entry_revenue")),
            ("Exit Multiple", _FMT_MULT, downside_result.get("exit_multiple")),
            ("IRR", _FMT_PCT, downside_result.get("irr")),
            ("MOIC", _FMT_MULT, downside_result.get("moic")),
        ]
        for label, fmt, value in downside_fields:
            ws.cell(row=r, column=1, value=f"  {label}").font = _FONT_NOTE
            vcell = ws.cell(row=r, column=2, value=value if value is not None else "N/A")
            if value is not None:
                vcell.number_format = fmt
            vcell.font = _FONT_NORMAL
            vcell.alignment = _ALIGN_RIGHT
            r += 1
        r += 1

    return r


def _build_assumptions_sheet(wb: Workbook, res: dict, lay: AssumptionsLayout, header_row: int):
    ws = wb[_SHEET_ASSUMPTIONS]

    _style_section_label(ws, header_row, 1, "OPERATING ASSUMPTIONS")
    _write_input(ws, lay.row_entry_revenue, 2, "Entry Revenue (Year 0)", res.get("entry_revenue", 0), _FMT_EUR)
    _write_input(ws, lay.row_ebitda_margin, 2, "EBITDA Margin", res.get("ebitda_margin", 0), _FMT_PCT)
    _write_input(ws, lay.row_revenue_growth, 2, "Revenue Growth (p.a.)", res.get("revenue_growth", 0), _FMT_PCT)
    _write_input(ws, lay.row_capex_pct, 2, "Capex (% Revenue)", res.get("capex_pct", 0), _FMT_PCT)
    _write_input(ws, lay.row_wcr_pct, 2, "ΔWCR (% ΔRevenue)", res.get("wcr_pct", 0), _FMT_PCT)
    _write_input(ws, lay.row_tax_rate, 2, "Tax Rate", res.get("tax_rate", 0), _FMT_PCT)
    _write_input(
        ws, lay.row_holding_period, 2, "Holding Period (yrs)", res.get("holding_period", 5),
        comment="Recomputes amortization pace, but does NOT add/remove projected year "
                "columns — the number of years is fixed at export time.",
    )
    _write_input(
        ws, lay.row_max_leverage, 2, "Max Leverage (% of EV)", MAX_LEVERAGE_PCT, _FMT_PCT,
        comment="Hard cap on total entry debt as % of Enterprise Value, exposed here as an "
                "editable input.",
    )
    _write_input(
        ws, lay.row_transaction_fee_pct, 2, "Transaction Fees (% of EV)",
        TRANSACTION_FEE_PCT, _FMT_PCT,
        comment="ESTIMATE — standard mid-market convention for M&A advisory/legal/DD fees "
                "in the absence of a real, negotiated fee letter. Added to Uses (Sources & "
                "Uses tab): Uses is no longer the Enterprise Value alone.",
    )
    _write_input(
        ws, lay.row_financing_fee_pct, 2, "Financing Fees (% of Debt Raised)",
        FINANCING_FEE_PCT, _FMT_PCT,
        comment="ESTIMATE — standard mid-market convention for arrangement/OID fees on the "
                "debt actually raised. Added to Uses.",
    )
    _write_input(
        ws, lay.row_min_cash_pct, 2, "Minimum Cash (% of Year-0 Revenue)",
        MIN_CASH_PCT, _FMT_PCT,
        comment="Minimum operating cash FUNDED AT CLOSING (Day-1 balance-sheet cash, not "
                "cash generated later) — added to Uses and carried as the Year-0 opening "
                "cash balance on the Debt Schedule tab.",
    )
    _write_input(
        ws, lay.row_dscr_covenant_min, 2, "DSCR Covenant (minimum)",
        DSCR_COVENANT_MIN, _FMT_MULT,
        comment="ESTIMATE — standard senior debt covenant threshold. Used by the Checks tab "
                "only; does not affect any cash-flow or debt formula.",
    )
    _write_input(
        ws, lay.row_leverage_covenant_stepdown, 2, "Leverage Covenant Step-Down (%/yr)",
        LEVERAGE_COVENANT_STEPDOWN_PCT, _FMT_PCT,
        comment="ESTIMATE — annual reduction applied to the entry net leverage to build a "
                "declining covenant cap (standard mid-market credit-agreement structure). "
                "Used by the Checks tab only.",
    )

    r = lay.row_leverage_covenant_stepdown + 2
    _style_section_label(ws, r, 1, "ENTRY / EXIT MULTIPLES")
    calibration = res.get("calibration") or {}
    if lay.has_calibration:
        tickers = ", ".join(calibration.get("tickers", []))
        years = calibration.get("fiscal_years", [])
        years_label = "-".join(str(y) for y in sorted(set(years))) if years else "N/A"
        _write_input(
            ws, lay.row_calib_median, 2, "Comparables Median EV/EBITDA",
            calibration.get("median_ev_ebitda", 0), _FMT_MULT,
            comment=(
                f"Median EV/EBITDA of CompSet '{calibration.get('comp_set_name')}' "
                f"(n={calibration.get('sample_size')}, FY{years_label}). Comparables: {tickers}."
            ),
        )
        _write_input(
            ws, lay.row_calib_discount, 2, "Size & Illiquidity Discount",
            calibration.get("size_illiquidity_discount", 0.35), _FMT_PCT,
            comment=calibration.get("discount_label", "Size & illiquidity discount, French mid-market"),
        )
        _write_formula(
            ws, lay.row_calib_derived, 2, "Derived Entry Multiple (calibrated)",
            f"=B{lay.row_calib_median}*(1-B{lay.row_calib_discount})", _FMT_MULT,
            fill_light=True,
        )
    _write_input(
        ws, lay.row_entry_multiple, 2, "Entry Multiple (used in model)", res.get("entry_multiple", 0), _FMT_MULT,
        comment=(
            "Defaults to the calibrated multiple above; edit freely to test an override."
            if lay.has_calibration else "EV/EBITDA multiple applied at entry."
        ),
    )
    _write_input(ws, lay.row_exit_multiple, 2, "Exit Multiple (used in model)", res.get("exit_multiple", 0), _FMT_MULT)

    r = lay.row_exit_multiple + 2
    _style_section_label(ws, r, 1, "SOURCES & USES (ENTRY)")
    _write_formula(
        ws, lay.row_entry_ebitda, 2, "Entry EBITDA",
        f"=B{lay.row_entry_revenue}*B{lay.row_ebitda_margin}", _FMT_EUR,
        comment="D&A (see Operating Model tab) is not modelled from a separate fixed-asset "
                "schedule — ESTIMATE convention: D&A = Capex (steady-state). It only feeds "
                "the tax shield (EBIT = EBITDA − D&A drives the tax base), never subtracted "
                "twice from cash flow.",
    )
    _write_formula(
        ws, lay.row_ev, 2, "Enterprise Value (EV)",
        f"=B{lay.row_entry_ebitda}*B{lay.row_entry_multiple}", _FMT_EUR, fill_light=True,
    )

    if lay.n_tranches:
        raw_col = _col(assum_layout_col(lay, "raw"))
        final_col = _col(assum_layout_col(lay, "final"))
        first_t, last_t = lay.tranche_rows[0], lay.tranche_rows[-1]
        raw_range = f"{raw_col}{first_t}:{raw_col}{last_t}"
        final_range = f"{final_col}{first_t}:{final_col}{last_t}"

        _write_formula(
            ws, lay.row_leverage_scale, 2, "Leverage Cap Scale Factor",
            f"=MIN(1,(B{lay.row_ev}*B{lay.row_max_leverage})/SUM({raw_range}))", _FMT_NUM_2,
        )
        _write_formula(
            ws, lay.row_total_debt, 2, "Total Debt (post leverage cap)",
            f"=SUM({final_range})", _FMT_EUR,
        )
    else:
        # V2 legacy: Total Debt is DERIVED (MIN(turns × EBITDA, EV × max leverage)),
        # exactly mirroring `_build_lbo`'s V2 branch — never copied as a rounded
        # snapshot of `entry_debt`, which would silently break both "flex an
        # assumption and watch it recompute" AND cent-level concordance with the
        # engine (a rounded starting balance would compound a small drift over
        # the Debt Schedule's multi-year simulation).
        _write_input(
            ws, lay.row_leverage_turns_v2, 2, "Leverage (Debt/EBITDA turns)",
            res.get("leverage_entry", 4.0), _FMT_MULT,
            comment="Turns of EBITDA applied to the senior debt tranche — defaults to a "
                    "standard 4.0x unless overridden. Total Debt below is capped at Max "
                    "Leverage (% of EV) above.",
        )
        _write_formula(
            ws, lay.row_total_debt, 2, "Total Debt (post leverage cap)",
            f"=MIN(B{lay.row_entry_ebitda}*B{lay.row_leverage_turns_v2},B{lay.row_ev}*B{lay.row_max_leverage})",
            _FMT_EUR,
        )
        _write_input(
            ws, lay.row_interest_rate_v2, 2, "Interest Rate (single tranche)",
            res.get("interest_rate", 0), _FMT_PCT,
            comment="Simple single-tranche debt structure: one senior debt block, fully "
                    "cash-swept every year (see Debt Schedule tab), not a fixed "
                    "amortization schedule.",
        )

    # Tâche "P1" (Partie E) — Uses réels : EV + frais de transaction (% EV) +
    # frais de financement (% de la dette EFFECTIVEMENT levée, donc après le
    # plafond de levier ci-dessus) + cash minimum financé à la clôture (%
    # du CA Année 0). L'equity sponsor (plug) absorbe ce total, pas l'EV
    # seul — Sources = Uses n'est donc plus une simple tautologie sur l'EV.
    _write_formula(
        ws, lay.row_transaction_fees, 2, "Transaction Fees",
        f"=B{lay.row_ev}*B{lay.row_transaction_fee_pct}", _FMT_EUR,
    )
    _write_formula(
        ws, lay.row_financing_fees, 2, "Financing Fees",
        f"=B{lay.row_total_debt}*B{lay.row_financing_fee_pct}", _FMT_EUR,
    )
    _write_formula(
        ws, lay.row_min_cash, 2, "Minimum Cash (funded at closing)",
        f"=B{lay.row_entry_revenue}*B{lay.row_min_cash_pct}", _FMT_EUR,
    )
    _write_formula(
        ws, lay.row_uses_total, 2, "Total Uses (EV + fees + minimum cash)",
        f"=B{lay.row_ev}+B{lay.row_transaction_fees}+B{lay.row_financing_fees}+B{lay.row_min_cash}",
        _FMT_EUR, fill_light=True,
    )
    _write_formula(
        ws, lay.row_sponsor_equity, 2, "Sponsor Equity (plug)",
        f"=B{lay.row_uses_total}-B{lay.row_total_debt}", _FMT_EUR, fill_light=True,
    )
    _write_formula(
        ws, lay.row_leverage_mult, 2, "Leverage (Debt/EBITDA)",
        f"=B{lay.row_total_debt}/B{lay.row_entry_ebitda}", _FMT_MULT,
    )

    if lay.n_tranches:
        r = lay.tranche_rows[0] - 2
        _style_section_label(ws, r, 1, "DEBT STRUCTURE (per tranche)")
        r += 1
        headers = ["Tranche", "Turns (x)", "Rate", "Amortization", "Raw Amount (turns×EBITDA)", "Final Amount (post cap)"]
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=r, column=c, value=h)
            cell.font = _FONT_HEADER
            cell.fill = _FILL_HEADER
            cell.alignment = _ALIGN_CENTER

        tranches = res.get("debt_tranches_detail", [])
        for i, t in enumerate(tranches):
            row = lay.tranche_rows[i]
            ws.cell(row=row, column=1, value=t["name"]).font = _FONT_NORMAL
            turns_cell = ws.cell(row=row, column=lay.COL_TURNS, value=t["turns"])
            turns_cell.font = _FONT_INPUT
            turns_cell.number_format = _FMT_MULT
            rate_cell = ws.cell(row=row, column=lay.COL_RATE, value=t["interest_rate"])
            rate_cell.font = _FONT_INPUT
            rate_cell.number_format = _FMT_PCT
            # Defensive coercion: `t["amortization"]` may be a str-Enum member
            # (AmortizationType) rather than a plain str, depending on the call
            # path (e.g. FastAPI's `.model_dump()` without mode="json" keeps the
            # enum instance). openpyxl calls str() internally on cell values,
            # and Enum.__str__ returns "AmortizationType.AMORTIZING" instead of
            # "amortizing" even though the member IS str-equal to it — silently
            # breaking the sheet's =EXACT() formulas (found live-testing the
            # /lbo/export-excel endpoint, RAPPORT B.9).
            amort_value = getattr(t["amortization"], "value", t["amortization"])
            amort_cell = ws.cell(row=row, column=lay.COL_AMORT_TYPE, value=amort_value)
            amort_cell.font = _FONT_INPUT
            amort_cell.comment = Comment(
                'Must be exactly "amortizing" or "bullet" (case-sensitive) — read live by '
                "the Debt Schedule sheet's =EXACT() formulas.", "PE Tracker",
            )

            raw_cell = ws.cell(
                row=row, column=lay.COL_RAW_AMOUNT,
                value=f"={_col(lay.COL_TURNS)}{row}*$B${lay.row_entry_ebitda}",
            )
            raw_cell.font = _FONT_NORMAL
            raw_cell.number_format = _FMT_EUR

            final_cell = ws.cell(
                row=row, column=lay.COL_FINAL_AMOUNT,
                value=f"={_col(lay.COL_RAW_AMOUNT)}{row}*$B${lay.row_leverage_scale}",
            )
            final_cell.font = _FONT_BOLD
            final_cell.number_format = _FMT_EUR

            if row % 2 == 0:
                for c in range(1, 7):
                    ws.cell(row=row, column=c).fill = _FILL_ALT

    _auto_width(ws)


def assum_layout_col(lay: AssumptionsLayout, which: str) -> int:
    return lay.COL_RAW_AMOUNT if which == "raw" else lay.COL_FINAL_AMOUNT


# ============================================================
# Sheet 2 — Sources & Uses (pure cross-sheet references — no duplicated input)
# ============================================================

def _build_sources_uses_sheet(wb: Workbook, res: dict, alay: AssumptionsLayout) -> dict:
    ws = wb.create_sheet(_SHEET_SOURCES_USES)
    ws.sheet_properties.tabColor = _DARK_BLUE
    _write_title(ws, "SOURCES & USES — entry (every cell links to Cover & Assumptions; nothing duplicated)", 3)

    r = 3
    _style_section_label(ws, r, 1, "USES")
    r += 1
    _write_formula(
        ws, r, 2, "Enterprise Value (Purchase of Target)",
        f"={_xref(_SHEET_ASSUMPTIONS, 2, alay.row_ev)}", _FMT_EUR, link=True,
    )
    r += 1
    _write_formula(
        ws, r, 2, "Transaction Fees (advisory/legal/DD)",
        f"={_xref(_SHEET_ASSUMPTIONS, 2, alay.row_transaction_fees)}", _FMT_EUR, bold=False, link=True,
    )
    r += 1
    _write_formula(
        ws, r, 2, "Financing Fees (debt arrangement)",
        f"={_xref(_SHEET_ASSUMPTIONS, 2, alay.row_financing_fees)}", _FMT_EUR, bold=False, link=True,
    )
    r += 1
    _write_formula(
        ws, r, 2, "Minimum Cash (funded at closing)",
        f"={_xref(_SHEET_ASSUMPTIONS, 2, alay.row_min_cash)}", _FMT_EUR, bold=False, link=True,
    )
    r += 1
    _write_formula(
        ws, r, 2, "Total Uses",
        f"={_xref(_SHEET_ASSUMPTIONS, 2, alay.row_uses_total)}", _FMT_EUR, fill_light=True, link=True,
    )
    uses_total_row = r
    r += 2

    _style_section_label(ws, r, 1, "SOURCES")
    r += 1
    source_rows: list[int] = []
    if alay.n_tranches:
        for t_row in alay.tranche_rows:
            name_cell = ws.cell(row=r, column=1, value=f"={_xref(_SHEET_ASSUMPTIONS, 1, t_row)}")
            name_cell.font = _FONT_LINK
            val_cell = ws.cell(
                row=r, column=2, value=f"={_xref(_SHEET_ASSUMPTIONS, alay.COL_FINAL_AMOUNT, t_row)}",
            )
            val_cell.font = _FONT_LINK_BOLD
            val_cell.number_format = _FMT_EUR
            val_cell.alignment = _ALIGN_RIGHT
            source_rows.append(r)
            r += 1
    else:
        ws.cell(row=r, column=1, value="Senior Debt (single tranche)").font = _FONT_NORMAL
        val_cell = ws.cell(row=r, column=2, value=f"={_xref(_SHEET_ASSUMPTIONS, 2, alay.row_total_debt)}")
        val_cell.font = _FONT_LINK_BOLD
        val_cell.number_format = _FMT_EUR
        val_cell.alignment = _ALIGN_RIGHT
        source_rows.append(r)
        r += 1

    _write_formula(
        ws, r, 2, "Sponsor Equity (plug)",
        f"={_xref(_SHEET_ASSUMPTIONS, 2, alay.row_sponsor_equity)}", _FMT_EUR, link=True,
    )
    source_rows.append(r)
    r += 1
    sources_sum = "+".join(f"B{row}" for row in source_rows)
    _write_formula(ws, r, 2, "Total Sources", f"={sources_sum}", _FMT_EUR, fill_light=True)
    sources_total_row = r
    r += 2

    # Tâche "P1" (Partie F) — ce n'est PLUS listé comme un "check" sur l'onglet
    # Checks : une equity plug équilibre TOUJOURS son propre total par
    # construction, quoi que Uses contienne — ce n'est donc pas un test qui
    # peut échouer (voir Checks tab). La ligne ci-dessous reste comme simple
    # rappel visuel de la mécanique Sources=Uses, pas comme un contrôle audité.
    _style_section_label(ws, r, 1, "Sources − Uses (always 0 by construction — see Checks tab for real tests)")
    r += 1
    _write_formula(
        ws, r, 2, "Sources − Uses",
        f"=ROUND(B{sources_total_row}-B{uses_total_row},2)", _FMT_EUR, bold=False,
    )
    diff_row = r
    r += 1

    _auto_width(ws)
    return {
        "uses_total_row": uses_total_row,
        "sources_total_row": sources_total_row,
        "diff_row": diff_row,
    }


# ============================================================
# Sheet 3 — Operating Model (P&L / Cash Flow Projections, 100% formula-driven)
# ============================================================

def _build_projections_sheet(
    wb: Workbook, res: dict, lay: PnlLayout, alay: AssumptionsLayout,
    dlay: DebtScheduleLayout | None, n_years: int,
):
    ws = wb.create_sheet(_SHEET_PNL)
    ws.sheet_properties.tabColor = _DARK_BLUE

    projs = res.get("projections", [])
    n_cols = n_years + 1
    _write_title(ws, "OPERATING MODEL — Revenue to Free Cash Flow (no hard-coded rates)", n_cols)

    hdr_row = 3
    for idx, p in enumerate(projs):
        ws.cell(row=hdr_row, column=idx + 2, value=f"Year {p['year']}").font = _FONT_HEADER
    _style_header_row(ws, hdr_row, n_cols)

    _style_section_label(ws, 4, 1, "P&L")
    ws.cell(row=4, column=1).comment = Comment(
        "D&A is not modelled from a separate fixed-asset schedule — ESTIMATE convention: "
        "D&A = Capex (steady-state, i.e. the asset base is replaced at the same pace it "
        "depreciates). It feeds ONLY the tax base (EBIT = EBITDA − D&A drives Taxable "
        "Income) — never subtracted a second time from cash flow, since Capex (a real "
        "cash outflow) already leaves the Free Cash Flow line below.", "PE Tracker",
    )
    for c in range(1, n_cols + 1):
        ws.cell(row=4, column=c).fill = _FILL_LIGHT

    A = alay  # shorthand

    def _assum(row: int) -> str:
        return _xref(_SHEET_ASSUMPTIONS, 2, row)

    # ── Revenue : Year0 = Assumptions!EntryRevenue, Year t = prior*(1+growth) ──
    r = lay.row_revenue
    ws.cell(row=r, column=1, value="Revenue").font = _FONT_BOLD
    for idx in range(len(projs)):
        col = idx + 2
        cell = ws.cell(row=r, column=col)
        if idx == 0:
            cell.value = f"={_assum(A.row_entry_revenue)}"
        else:
            cell.value = f"={_col(col - 1)}{r}*(1+{_xref(_SHEET_ASSUMPTIONS, 2, A.row_revenue_growth)})"
        cell.number_format = _FMT_EUR
        cell.font = _FONT_BOLD
        cell.alignment = _ALIGN_RIGHT

    # ── EBITDA = Revenue × Margin (every year, incl. Year 0) ──
    r = lay.row_ebitda
    ws.cell(row=r, column=1, value="EBITDA").font = _FONT_BOLD
    for idx in range(len(projs)):
        col = idx + 2
        cell = ws.cell(row=r, column=col,
                        value=f"={_col(col)}{lay.row_revenue}*{_xref(_SHEET_ASSUMPTIONS, 2, A.row_ebitda_margin)}")
        cell.number_format = _FMT_EUR
        cell.font = _FONT_BOLD
        cell.alignment = _ALIGN_RIGHT

    # ── D&A = Capex (steady-state ESTIMATE convention — see row-4 comment) ──
    # Forward-references the Capex row below (Excel resolves by dependency
    # graph, not by row order — no circularity: Capex depends only on
    # Revenue/Assumptions, D&A depends only on Capex).
    r = lay.row_da
    ws.cell(row=r, column=1, value="(-) D&A (= Capex, steady-state ESTIMATE)").font = _FONT_NORMAL
    for idx in range(len(projs)):
        col = idx + 2
        if idx == 0:
            cell = ws.cell(row=r, column=col, value=0)
        else:
            cell = ws.cell(row=r, column=col, value=f"={_col(col)}{lay.row_capex}")
        cell.number_format = _FMT_EUR
        cell.alignment = _ALIGN_RIGHT

    # ── EBIT = EBITDA - D&A (Partie C — feeds the tax shield below) ──
    r = lay.row_ebit
    ws.cell(row=r, column=1, value="EBIT (EBITDA − D&A)").font = _FONT_BOLD
    for idx in range(len(projs)):
        col = idx + 2
        if idx == 0:
            cell = ws.cell(row=r, column=col, value=0)
        else:
            cell = ws.cell(row=r, column=col,
                            value=f"={_col(col)}{lay.row_ebitda}-{_col(col)}{lay.row_da}")
        cell.number_format = _FMT_EUR
        cell.font = _FONT_BOLD
        cell.alignment = _ALIGN_RIGHT

    # ── Interest = SUM of per-tranche interest, from Debt Schedule (cross-sheet) ──
    r = lay.row_interest
    ws.cell(row=r, column=1, value="(-) Interest Expense").font = _FONT_NORMAL
    for idx in range(len(projs)):
        col = idx + 2
        if dlay:
            cell = ws.cell(row=r, column=col, value=f"={_xref(_SHEET_DEBT, col, dlay.row_total_interest)}")
            cell.font = _FONT_LINK
        else:
            cell = ws.cell(row=r, column=col, value=projs[idx].get("interest", 0))
            cell.font = _FONT_NORMAL
        cell.number_format = _FMT_EUR
        cell.alignment = _ALIGN_RIGHT

    # ── Taxable Income (EBT) = EBIT − Interest (Partie C: D&A tax shield) ──
    # Year 0 is the entry point, not an operating year — the engine's own
    # projections[0] hard-zeros taxable_income/tax/capex/fcf (only revenue,
    # ebitda and debt_eoy are non-zero at Year 0). Mirrored here explicitly;
    # omitting the idx==0 branch would show a fabricated Year-0 tax/capex/FCF
    # the engine never computes (found and fixed during verification).
    r = lay.row_taxable_income
    ws.cell(row=r, column=1, value="Taxable Income (EBT = EBIT − Interest)").font = _FONT_NORMAL
    for idx in range(len(projs)):
        col = idx + 2
        if idx == 0:
            cell = ws.cell(row=r, column=col, value=0)
        else:
            cell = ws.cell(row=r, column=col,
                            value=f"={_col(col)}{lay.row_ebit}-{_col(col)}{lay.row_interest}")
        cell.number_format = _FMT_EUR
        cell.alignment = _ALIGN_RIGHT

    # ── Tax = MAX(0, EBT × tax_rate) — 0 at Year 0, see note above ──
    r = lay.row_tax
    ws.cell(row=r, column=1, value="(-) Corporate Tax").font = _FONT_NORMAL
    for idx in range(len(projs)):
        col = idx + 2
        if idx == 0:
            cell = ws.cell(row=r, column=col, value=0)
        else:
            cell = ws.cell(
                row=r, column=col,
                value=f"=MAX(0,{_col(col)}{lay.row_taxable_income}*{_xref(_SHEET_ASSUMPTIONS, 2, A.row_tax_rate)})",
            )
        cell.number_format = _FMT_EUR
        cell.alignment = _ALIGN_RIGHT

    r0 = lay.row_tax + 1
    _style_section_label(ws, r0, 1, "CASH FLOW")
    for c in range(1, n_cols + 1):
        ws.cell(row=r0, column=c).fill = _FILL_LIGHT

    # ── Capex = Revenue × capex_pct — 0 at Year 0, see note above ──
    r = lay.row_capex
    ws.cell(row=r, column=1, value="(-) Capex").font = _FONT_NORMAL
    for idx in range(len(projs)):
        col = idx + 2
        if idx == 0:
            cell = ws.cell(row=r, column=col, value=0)
        else:
            cell = ws.cell(
                row=r, column=col,
                value=f"={_col(col)}{lay.row_revenue}*{_xref(_SHEET_ASSUMPTIONS, 2, A.row_capex_pct)}",
            )
        cell.number_format = _FMT_EUR
        cell.alignment = _ALIGN_RIGHT

    # ── ΔWCR = MAX(0,(Rev_t - Rev_{t-1}) × wcr_pct) ──
    r = lay.row_wcr
    ws.cell(row=r, column=1, value="(-) ΔWCR").font = _FONT_NORMAL
    for idx in range(len(projs)):
        col = idx + 2
        if idx == 0:
            cell = ws.cell(row=r, column=col, value=0)
        else:
            cell = ws.cell(
                row=r, column=col,
                value=(
                    f"=MAX(0,({_col(col)}{lay.row_revenue}-{_col(col-1)}{lay.row_revenue})"
                    f"*{_xref(_SHEET_ASSUMPTIONS, 2, A.row_wcr_pct)})"
                ),
            )
        cell.number_format = _FMT_EUR
        cell.alignment = _ALIGN_RIGHT

    # ── FCF = EBITDA - Interest - Tax - Capex - ΔWCR ──
    r = lay.row_fcf
    ws.cell(row=r, column=1, value="Free Cash Flow").font = _FONT_BOLD
    for idx in range(len(projs)):
        col = idx + 2
        if idx == 0:
            # Year 0 is the entry date itself — no operating cash flow happens
            # ON the transaction date in this engine (`projections[0]["fcf"] = 0.0`
            # explicitly); all other Year-0 P&L lines are already 0 above, so
            # this would otherwise wrongly evaluate to +EBITDA.
            cell = ws.cell(row=r, column=col, value=0)
        else:
            cell = ws.cell(
                row=r, column=col,
                value=(
                    f"={_col(col)}{lay.row_ebitda}-{_col(col)}{lay.row_interest}"
                    f"-{_col(col)}{lay.row_tax}-{_col(col)}{lay.row_capex}-{_col(col)}{lay.row_wcr}"
                ),
            )
        cell.number_format = _FMT_EUR
        cell.font = _FONT_BOLD
        cell.alignment = _ALIGN_RIGHT

    r1 = lay.row_fcf + 1
    _style_section_label(ws, r1, 1, "DEBT SCHEDULE (see Debt Schedule tab for detail)")
    for c in range(1, n_cols + 1):
        ws.cell(row=r1, column=c).fill = _FILL_LIGHT

    # ── Debt Repayment / Debt Outstanding — cross-sheet references ──
    r = lay.row_debt_repayment
    ws.cell(row=r, column=1, value="Debt Repayment").font = _FONT_NORMAL
    for idx in range(len(projs)):
        col = idx + 2
        if dlay:
            cell = ws.cell(row=r, column=col, value=f"={_xref(_SHEET_DEBT, col, dlay.row_total_repayment)}")
            cell.font = _FONT_LINK
        else:
            cell = ws.cell(row=r, column=col, value=projs[idx].get("debt_paydown", 0))
            cell.font = _FONT_NORMAL
        cell.number_format = _FMT_EUR
        cell.alignment = _ALIGN_RIGHT

    r = lay.row_debt_eoy
    ws.cell(row=r, column=1, value="Debt Outstanding (EOY, gross)").font = _FONT_BOLD
    for idx in range(len(projs)):
        col = idx + 2
        if dlay:
            cell = ws.cell(row=r, column=col, value=f"={_xref(_SHEET_DEBT, col, dlay.row_total_debt_eoy)}")
            cell.font = _FONT_LINK_BOLD
        else:
            cell = ws.cell(row=r, column=col, value=projs[idx].get("debt_eoy", 0))
            cell.font = _FONT_BOLD
        cell.number_format = _FMT_EUR
        cell.alignment = _ALIGN_RIGHT

    # ── Cash / Net Debt (Partie B) — the cash sweep no longer discards
    # unabsorbed FCF: it accumulates on the Debt Schedule tab and is netted
    # against gross debt here, exactly what Exit Equity uses on the Returns tab.
    r = lay.row_cash_eoy
    ws.cell(row=r, column=1, value="Cash Balance (EOY)").font = _FONT_NORMAL
    for idx in range(len(projs)):
        col = idx + 2
        if dlay:
            cell = ws.cell(row=r, column=col, value=f"={_xref(_SHEET_DEBT, col, dlay.row_cash_balance_eoy)}")
            cell.font = _FONT_LINK
        else:
            cell = ws.cell(row=r, column=col, value=projs[idx].get("cash_eoy", 0))
            cell.font = _FONT_NORMAL
        cell.number_format = _FMT_EUR
        cell.alignment = _ALIGN_RIGHT

    r = lay.row_net_debt_eoy
    ws.cell(row=r, column=1, value="Net Debt (EOY)").font = _FONT_BOLD
    for idx in range(len(projs)):
        col = idx + 2
        if dlay:
            cell = ws.cell(row=r, column=col, value=f"={_xref(_SHEET_DEBT, col, dlay.row_net_debt_eoy)}")
            cell.font = _FONT_LINK_BOLD
        else:
            cell = ws.cell(row=r, column=col, value=projs[idx].get("net_debt_eoy", 0))
            cell.font = _FONT_BOLD
        cell.number_format = _FMT_EUR
        cell.alignment = _ALIGN_RIGHT
        cell.border = _THICK_BOTTOM

    _auto_width(ws, min_width=14)


# ============================================================
# Sheet 4 — Debt Schedule (per-tranche OR V2 single cash-swept tranche)
# ============================================================

def _build_debt_schedule_sheet(
    wb: Workbook, res: dict, lay: DebtScheduleLayout, alay: AssumptionsLayout,
    play: PnlLayout, n_years: int, *, is_v2: bool = False,
):
    ws = wb.create_sheet(_SHEET_DEBT)
    ws.sheet_properties.tabColor = _GOLD

    projs = res.get("projections", [])
    n_cols = n_years + 1

    if is_v2:
        title = "DEBT SCHEDULE — single senior tranche, contractual schedule + cash sweep"
        tranches_meta = [{"name": "Senior Debt"}]
    else:
        title = "DEBT SCHEDULE — fully formulaic, per-tranche amortization + cash sweep"
        tranches_meta = res.get("debt_tranches_detail", [])

    _write_title(ws, title, n_cols)

    hdr_row = 3
    for idx, p in enumerate(projs):
        ws.cell(row=hdr_row, column=idx + 2, value=f"Year {p['year']}").font = _FONT_HEADER
    _style_header_row(ws, hdr_row, n_cols)

    A = alay
    holding_ref = _xref(_SHEET_ASSUMPTIONS, 2, A.row_holding_period)
    min_cash_ref = _xref(_SHEET_ASSUMPTIONS, 2, A.row_min_cash)
    last_block = lay.tranche_blocks[-1]

    for t_idx, (t_meta, block) in enumerate(zip(tranches_meta, lay.tranche_blocks)):
        t_name = t_meta["name"]
        if is_v2:
            rate_ref = _xref(_SHEET_ASSUMPTIONS, 2, A.row_interest_rate_v2)
            final_amount_ref = _xref(_SHEET_ASSUMPTIONS, 2, A.row_total_debt)
            amort_type_ref = None
        else:
            assum_row = A.tranche_rows[t_idx]
            rate_ref = _xref(_SHEET_ASSUMPTIONS, A.COL_RATE, assum_row)
            amort_type_ref = _xref(_SHEET_ASSUMPTIONS, A.COL_AMORT_TYPE, assum_row)
            final_amount_ref = _xref(_SHEET_ASSUMPTIONS, A.COL_FINAL_AMOUNT, assum_row)

        _style_section_label(ws, block["cash_avail_sched"] - 1, 1, t_name.upper())
        for c in range(1, n_cols + 1):
            ws.cell(row=block["cash_avail_sched"] - 1, column=c).fill = _FILL_LIGHT

        ws.cell(row=block["cash_avail_sched"], column=1, value="  Cash Available for Scheduled Amort. (cascade)").font = _FONT_NOTE
        ws.cell(row=block["opening"], column=1, value="  Opening Balance").font = _FONT_NORMAL
        ws.cell(row=block["interest"], column=1, value="  Interest").font = _FONT_NORMAL
        ws.cell(row=block["sched_due"], column=1, value="  Scheduled Amort. DUE (contractual)").font = _FONT_NORMAL
        ws.cell(row=block["sched_paid"], column=1, value="  Scheduled Amort. PAID (cash-capped)").font = _FONT_NORMAL
        ws.cell(row=block["cash_avail_sweep"], column=1, value="  Cash Available for Sweep (cascade)").font = _FONT_NOTE
        ws.cell(row=block["sweep_repay"], column=1, value="  Cash Sweep Repayment").font = _FONT_NORMAL
        ws.cell(row=block["total_repay"], column=1, value="  Total Repayment (scheduled + sweep)").font = _FONT_NORMAL
        ws.cell(row=block["closing"], column=1, value="  Closing Balance").font = _FONT_BOLD

        for idx, p in enumerate(projs):
            col = idx + 2

            # Cash available for THIS tranche's scheduled amortization —
            # cascaded by seniority (tranche order = seniority order): the
            # most senior tranche gets first claim on the year's FCF, each
            # junior tranche gets what the senior one(s) didn't need.
            cash_sched_cell = ws.cell(row=block["cash_avail_sched"], column=col)
            if idx == 0:
                cash_sched_cell.value = 0
            elif t_idx == 0:
                cash_sched_cell.value = f"=MAX(0,{_xref(_SHEET_PNL, col, play.row_fcf)})"
            else:
                prev = lay.tranche_blocks[t_idx - 1]
                cash_sched_cell.value = f"={_col(col)}{prev['cash_avail_sched']}-{_col(col)}{prev['sched_paid']}"
            cash_sched_cell.number_format = _FMT_EUR
            cash_sched_cell.font = _FONT_NOTE

            # Opening balance
            open_cell = ws.cell(row=block["opening"], column=col)
            if idx == 0:
                open_cell.value = f"={final_amount_ref}"
                open_cell.font = _FONT_LINK
            else:
                open_cell.value = f"={_col(col-1)}{block['closing']}"
                open_cell.font = _FONT_NORMAL
            open_cell.number_format = _FMT_EUR

            # Interest = opening balance × rate (no interest accrues at Year 0 funding).
            # Tâche "P1" (Partie B) — reste calculé sur le solde d'OUVERTURE
            # (pas de clôture) pour éviter la circularité intérêt↔FCF↔
            # remboursement↔solde de clôture ; choix déjà en place, conservé
            # et documenté ici comme demandé.
            int_cell = ws.cell(row=block["interest"], column=col)
            if idx == 0:
                int_cell.value = 0
            else:
                int_cell.value = f"={_col(col)}{block['opening']}*{rate_ref}"
            int_cell.number_format = _FMT_EUR
            int_cell.font = _FONT_NORMAL

            # Scheduled Amort. DUE — CONTRACTUAL amount owed this year,
            # capped only by the remaining balance, NEVER by cash
            # availability (Erreur 2 du rapport IC : avant, un simple MIN()
            # confondait "dû" et "payé", si bien qu'un manquement de cash
            # était silencieusement absorbé sans qu'aucun check ne le voie).
            # V2 legacy has no fixed schedule (annual_amort=0 in the engine)
            # — DUE is always 0, matching `_build_lbo`'s V2 branch exactly.
            due_cell = ws.cell(row=block["sched_due"], column=col)
            if idx == 0:
                due_cell.value = 0
            elif is_v2:
                due_cell.value = 0
            else:
                fixed_annuity = f"({final_amount_ref}/{holding_ref})"
                due_cell.value = (
                    f'=IF(EXACT({amort_type_ref},"amortizing"),'
                    f"MIN({fixed_annuity},{_col(col)}{block['opening']}),0)"
                )
            due_cell.number_format = _FMT_EUR
            due_cell.font = _FONT_NORMAL

            # Scheduled Amort. PAID — the DUE amount, capped by cash actually
            # available. DUE − PAID (summed across tranches) is the year's
            # debt-service shortfall — never silently absorbed, see Checks tab.
            paid_cell = ws.cell(row=block["sched_paid"], column=col)
            if idx == 0:
                paid_cell.value = 0
            else:
                paid_cell.value = f"=MIN({_col(col)}{block['sched_due']},{_col(col)}{block['cash_avail_sched']})"
            paid_cell.number_format = _FMT_EUR
            paid_cell.font = _FONT_NORMAL

            # Cash available for the SWEEP pass — a SECOND cascade, by the
            # same seniority order, but only starting once ALL tranches'
            # scheduled amortization has been settled (hence sourced from the
            # LAST tranche's post-scheduled leftover, not this tranche's own
            # pre-scheduled cash). Tâche "P1" (Partie B) — this pass now
            # applies UNIFORMLY to every tranche, bullet included: a 100%
            # cash sweep provision does not exempt an in-fine tranche from
            # early paydown, it only exempts it from a FIXED schedule
            # (before this fix, V3 bullets never received any sweep at all —
            # Erreur 1 du rapport IC — and V2's single tranche, once fully
            # repaid, let any further FCF evaporate with no cash line to
            # catch it).
            sweep_cash_cell = ws.cell(row=block["cash_avail_sweep"], column=col)
            if idx == 0:
                sweep_cash_cell.value = 0
            elif t_idx == 0:
                sweep_cash_cell.value = (
                    f"={_col(col)}{last_block['cash_avail_sched']}-{_col(col)}{last_block['sched_paid']}"
                )
            else:
                prev = lay.tranche_blocks[t_idx - 1]
                sweep_cash_cell.value = f"={_col(col)}{prev['cash_avail_sweep']}-{_col(col)}{prev['sweep_repay']}"
            sweep_cash_cell.number_format = _FMT_EUR
            sweep_cash_cell.font = _FONT_NOTE

            sweep_cell = ws.cell(row=block["sweep_repay"], column=col)
            if idx == 0:
                sweep_cell.value = 0
            else:
                sweep_cell.value = (
                    f"=MIN({_col(col)}{block['opening']}-{_col(col)}{block['sched_paid']},"
                    f"{_col(col)}{block['cash_avail_sweep']})"
                )
            sweep_cell.number_format = _FMT_EUR
            sweep_cell.font = _FONT_NORMAL

            total_repay_cell = ws.cell(row=block["total_repay"], column=col)
            total_repay_cell.value = f"={_col(col)}{block['sched_paid']}+{_col(col)}{block['sweep_repay']}"
            total_repay_cell.number_format = _FMT_EUR
            total_repay_cell.font = _FONT_BOLD

            # Closing balance
            close_cell = ws.cell(row=block["closing"], column=col)
            close_cell.value = f"={_col(col)}{block['opening']}-{_col(col)}{block['total_repay']}"
            close_cell.number_format = _FMT_EUR
            close_cell.font = _FONT_BOLD
            if idx == len(projs) - 1:
                close_cell.fill = _FILL_LIGHT

    # ── Totals ──
    r = lay.row_total_debt_eoy
    ws.cell(row=r, column=1, value="TOTAL DEBT EOY (gross)").font = _FONT_SECTION
    ws.cell(row=r, column=1).fill = _FILL_LIGHT
    for idx in range(len(projs)):
        col = idx + 2
        closing_refs = "+".join(f"{_col(col)}{b['closing']}" for b in lay.tranche_blocks)
        cell = ws.cell(row=r, column=col, value=f"={closing_refs}")
        cell.number_format = _FMT_EUR
        cell.font = _FONT_BOLD
        cell.fill = _FILL_LIGHT

    r = lay.row_total_interest
    ws.cell(row=r, column=1, value="TOTAL INTEREST").font = _FONT_SECTION
    for idx in range(len(projs)):
        col = idx + 2
        interest_refs = "+".join(f"{_col(col)}{b['interest']}" for b in lay.tranche_blocks)
        cell = ws.cell(row=r, column=col, value=f"={interest_refs}")
        cell.number_format = _FMT_EUR
        cell.font = _FONT_BOLD

    r = lay.row_total_sched_due
    ws.cell(row=r, column=1, value="TOTAL SCHEDULED AMORT. DUE (contractual)").font = _FONT_SECTION
    for idx in range(len(projs)):
        col = idx + 2
        refs = "+".join(f"{_col(col)}{b['sched_due']}" for b in lay.tranche_blocks)
        cell = ws.cell(row=r, column=col, value=f"={refs}")
        cell.number_format = _FMT_EUR
        cell.font = _FONT_BOLD

    r = lay.row_total_sched_paid
    ws.cell(row=r, column=1, value="TOTAL SCHEDULED AMORT. PAID").font = _FONT_SECTION
    for idx in range(len(projs)):
        col = idx + 2
        refs = "+".join(f"{_col(col)}{b['sched_paid']}" for b in lay.tranche_blocks)
        cell = ws.cell(row=r, column=col, value=f"={refs}")
        cell.number_format = _FMT_EUR
        cell.font = _FONT_BOLD

    r = lay.row_total_shortfall
    ws.cell(row=r, column=1, value="DEBT SERVICE SHORTFALL (DUE − PAID)").font = _FONT_SECTION
    ws.cell(row=r, column=1).comment = Comment(
        "Should be 0 every year — a positive value means the year's cash could not cover "
        "the contractually scheduled amortization (a technical default/covenant tension, "
        "never silently absorbed — see the Checks tab).", "PE Tracker",
    )
    for idx in range(len(projs)):
        col = idx + 2
        cell = ws.cell(
            row=r, column=col,
            value=f"=ROUND({_col(col)}{lay.row_total_sched_due}-{_col(col)}{lay.row_total_sched_paid},2)",
        )
        cell.number_format = _FMT_EUR
        cell.font = _FONT_BOLD

    r = lay.row_total_repayment
    ws.cell(row=r, column=1, value="TOTAL REPAYMENT (scheduled + sweep)").font = _FONT_SECTION
    for idx in range(len(projs)):
        col = idx + 2
        repay_refs = "+".join(f"{_col(col)}{b['total_repay']}" for b in lay.tranche_blocks)
        cell = ws.cell(row=r, column=col, value=f"={repay_refs}")
        cell.number_format = _FMT_EUR
        cell.font = _FONT_BOLD

    # ── Cash Balance & Net Debt (Partie B) — cash the sweep leaves unabsorbed
    # once every tranche is fully repaid no longer evaporates: it accumulates
    # here, starting from the minimum cash funded at closing (Year 0).
    r = lay.row_cash_balance_eoy
    ws.cell(row=r, column=1, value="CASH BALANCE (EOY)").font = _FONT_SECTION
    ws.cell(row=r, column=1).fill = _FILL_LIGHT
    for idx in range(len(projs)):
        col = idx + 2
        if idx == 0:
            cell = ws.cell(row=r, column=col, value=f"={min_cash_ref}")
        else:
            leftover = f"{_col(col)}{last_block['cash_avail_sweep']}-{_col(col)}{last_block['sweep_repay']}"
            cell = ws.cell(row=r, column=col, value=f"={_col(col-1)}{r}+{leftover}")
        cell.number_format = _FMT_EUR
        cell.font = _FONT_BOLD
        cell.fill = _FILL_LIGHT

    r = lay.row_net_debt_eoy
    ws.cell(row=r, column=1, value="NET DEBT (EOY) = Total Debt − Cash").font = _FONT_SECTION
    ws.cell(row=r, column=1).fill = _FILL_LIGHT
    for idx in range(len(projs)):
        col = idx + 2
        cell = ws.cell(
            row=r, column=col,
            value=f"={_col(col)}{lay.row_total_debt_eoy}-{_col(col)}{lay.row_cash_balance_eoy}",
        )
        cell.number_format = _FMT_EUR
        cell.font = _FONT_BOLD
        cell.fill = _FILL_LIGHT

    _auto_width(ws, min_width=14)


# ============================================================
# Sheet 5 — Returns & Waterfall (Exit EV / MoIC / IRR — all formulas)
# ============================================================

def _build_returns_sheet(
    wb: Workbook, res: dict, alay: AssumptionsLayout, play: PnlLayout | None,
    dlay: DebtScheduleLayout | None, n_years: int,
) -> dict:
    ws = wb.create_sheet(_SHEET_RETURNS)
    ws.sheet_properties.tabColor = _GREEN
    _write_title(ws, "RETURNS & EXIT ANALYSIS", 4)

    A = alay
    last_col = n_years + 1  # last projection year's column on P&L / Debt Schedule sheets

    r = 3
    _style_section_label(ws, r, 1, "EXIT ASSUMPTIONS")
    r += 1
    _write_formula(ws, r, 2, "Holding Period", f"={_xref(_SHEET_ASSUMPTIONS, 2, A.row_holding_period)}", bold=False, link=True)
    r += 1
    if play is not None:
        # ROUND(...,2) mirrors the engine exactly: `exit_ebitda = exit_year["ebitda"]`
        # reuses the ALREADY round2()-ed value from the projections dict (valuation_engine.py
        # _build_lbo), not the raw unrounded float — without this, Exit EV would drift by a
        # few cents against the engine (observed and traced during verification, RAPPORT B.9).
        exit_ebitda_formula = f"=ROUND({_xref(_SHEET_PNL, last_col, play.row_ebitda)},2)"
    else:
        exit_ebitda_formula = res.get("exit_ebitda", 0)
    _write_formula(ws, r, 2, "Exit EBITDA", exit_ebitda_formula, _FMT_EUR, bold=False, link=(play is not None))
    exit_ebitda_row = r
    r += 1
    _write_formula(ws, r, 2, "Exit Multiple", f"={_xref(_SHEET_ASSUMPTIONS, 2, A.row_exit_multiple)}", _FMT_MULT, bold=False, link=True)
    exit_mult_row = r
    r += 1

    _write_formula(ws, r, 2, "Exit EV", f"=B{exit_ebitda_row}*B{exit_mult_row}", _FMT_EUR, fill_light=True)
    exit_ev_row = r
    r += 1

    if dlay is not None:
        # Same ROUND(...,2) rationale as Exit EBITDA above — engine reuses
        # `exit_year["debt_eoy"]`, already rounded in the projections dict.
        exit_debt_formula = f"=ROUND({_xref(_SHEET_DEBT, last_col, dlay.row_total_debt_eoy)},2)"
        exit_cash_formula = f"=ROUND({_xref(_SHEET_DEBT, last_col, dlay.row_cash_balance_eoy)},2)"
    else:
        exit_debt_formula = res.get("exit_debt", 0)
        exit_cash_formula = res.get("exit_cash", 0)
    _write_formula(ws, r, 2, "Remaining Debt (gross)", exit_debt_formula, _FMT_EUR, bold=False, link=(dlay is not None))
    exit_debt_row = r
    r += 1
    _write_formula(ws, r, 2, "Remaining Cash", exit_cash_formula, _FMT_EUR, bold=False, link=(dlay is not None))
    exit_cash_row = r
    r += 1
    _write_formula(
        ws, r, 2, "Net Debt (gross debt − cash)", f"=B{exit_debt_row}-B{exit_cash_row}", _FMT_EUR, bold=False,
    )
    exit_net_debt_row = r
    r += 1

    # Tâche "P1" (Partie B) — Exit Equity nette de trésorerie : le cash
    # accumulé par le sweep (au lieu de s'évaporer) appartient au sponsor à
    # la sortie exactement comme une dette évitée.
    _write_formula(ws, r, 2, "Exit Equity", f"=B{exit_ev_row}-B{exit_net_debt_row}", _FMT_EUR, fill_light=True)
    exit_eq_row = r
    r += 2

    _style_section_label(ws, r, 1, "GROSS RETURNS (PRE-WATERFALL)")
    r += 1
    _write_formula(
        ws, r, 2, "Entry Equity", f"={_xref(_SHEET_ASSUMPTIONS, 2, A.row_sponsor_equity)}", _FMT_EUR, bold=False, link=True,
    )
    entry_eq_row = r
    r += 1

    _write_formula(ws, r, 2, "MoIC (Gross)", f"=B{exit_eq_row}/B{entry_eq_row}", _FMT_MULT, kpi=True)
    moic_row = r
    r += 1

    n = res.get("holding_period", 5)
    ws.cell(row=r, column=1, value="IRR (Gross) — native =IRR()").font = _FONT_NORMAL
    irr_cell = ws.cell(row=r, column=2)
    irr_row = r

    cf_start_row = r + 3
    ws.cell(row=cf_start_row - 1, column=1, value="Cash Flows (helper, referenced by =IRR() above):").font = _FONT_NOTE
    for i in range(n + 1):
        cf_row = cf_start_row + i
        cell = ws.cell(row=cf_row, column=1)
        if i == 0:
            cell.value = f"=-B{entry_eq_row}"
        elif i == n:
            cell.value = f"=B{exit_eq_row}"
        else:
            cell.value = 0
        cell.number_format = _FMT_EUR
        cell.font = _FONT_NOTE

    cf_range = f"A{cf_start_row}:A{cf_start_row + n}"
    irr_cell.value = f"=IRR({cf_range})"
    irr_cell.font = _FONT_KPI
    irr_cell.number_format = _FMT_PCT
    r += 1

    layout = {
        "entry_eq_row": entry_eq_row, "exit_eq_row": exit_eq_row, "exit_ev_row": exit_ev_row,
        "exit_debt_row": exit_debt_row, "exit_cash_row": exit_cash_row,
        "exit_net_debt_row": exit_net_debt_row, "moic_row": moic_row, "irr_row": irr_row,
        "holding_period": n,
    }

    # ── Waterfall (Fund vs Management) — formula-driven from two blue inputs ──
    wf = res.get("waterfall")
    if wf:
        r = cf_start_row + n + 2
        _style_section_label(ws, r, 1, "EQUITY WATERFALL — FUND vs MANAGEMENT")
        r += 1

        _write_input(
            ws, r, 2, "Mgmt Sweet Equity %", wf.get("management_sweet_pct", 0), _FMT_PCT,
        )
        sweet_pct_row = r
        r += 1
        _write_input(
            ws, r, 2, "Mgmt Total % (post-ratchet)", wf.get("management_total_pct", 0), _FMT_PCT,
            comment=(
                "Input, not recomputed live: the ratchet IRR threshold and bonus % are not "
                "exposed by the Python engine's output payload, so the ratchet decision "
                "cannot be re-evaluated inside Excel. This reflects the engine's decision "
                f"at generation time (ratchet_triggered={wf.get('ratchet_triggered')}). "
                "Editing IRR-driving assumptions above will NOT re-trigger/un-trigger it."
            ),
        )
        mgmt_pct_row = r
        r += 1
        ws.cell(row=r, column=1, value="Ratchet Triggered? (at generation time)").font = _FONT_NOTE
        ws.cell(row=r, column=2, value="YES" if wf.get("ratchet_triggered") else "NO").font = _FONT_NOTE
        r += 2

        _write_formula(ws, r, 2, "Total Exit Equity", f"=B{exit_eq_row}", _FMT_EUR)
        total_eq_row = r
        r += 1
        _write_formula(ws, r, 2, "Management Proceeds", f"=B{total_eq_row}*B{mgmt_pct_row}", _FMT_EUR)
        mgmt_proceeds_row = r
        r += 1
        _write_formula(ws, r, 2, "Fund Proceeds", f"=B{total_eq_row}-B{mgmt_proceeds_row}", _FMT_EUR)
        fund_proceeds_row = r
        r += 1

        # Tâche "P0 : un seul deal dans les 3 documents" (Partie C) : le
        # management co-investit sa sweet equity — le fonds n'apporte donc
        # PAS 100% de l'entry equity. "Mgmt Invested" doit être connu AVANT
        # "Fund Invested"/"Fund MoIC" (précédemment calculé après, avec Fund
        # MoIC rapporté à tort à l'entry equity complet — sous-évaluait le
        # MOIC fonds d'environ 1/(1−sweet_pct)).
        _write_formula(
            ws, r, 2, "Mgmt Invested (Entry Equity × Sweet %)",
            f"=B{entry_eq_row}*B{sweet_pct_row}", _FMT_EUR, bold=False,
        )
        mgmt_invested_row = r
        r += 1
        _write_formula(
            ws, r, 2, "Fund Invested (Entry Equity − Mgmt Invested)",
            f"=B{entry_eq_row}-B{mgmt_invested_row}", _FMT_EUR, bold=False,
        )
        fund_invested_row = r
        r += 1

        _write_formula(ws, r, 2, "Fund MoIC", f"=B{fund_proceeds_row}/B{fund_invested_row}", _FMT_MULT, kpi=True)
        r += 1

        ws.cell(row=r, column=1, value="Fund IRR — native =IRR()").font = _FONT_NORMAL
        fund_irr_cell = ws.cell(row=r, column=2)
        fund_cf_start = r + 3
        ws.cell(row=fund_cf_start - 1, column=1, value="Fund Cash Flows (helper):").font = _FONT_NOTE
        for i in range(n + 1):
            cf_row = fund_cf_start + i
            cell = ws.cell(row=cf_row, column=1)
            if i == 0:
                cell.value = f"=-B{fund_invested_row}"
            elif i == n:
                cell.value = f"=B{fund_proceeds_row}"
            else:
                cell.value = 0
            cell.number_format = _FMT_EUR
            cell.font = _FONT_NOTE
        fund_irr_cell.value = f"=IRR(A{fund_cf_start}:A{fund_cf_start + n})"
        fund_irr_cell.font = _FONT_KPI
        fund_irr_cell.number_format = _FMT_PCT
        r = fund_cf_start + n + 1

        _write_formula(ws, r, 2, "Management MoIC", f"=B{mgmt_proceeds_row}/B{mgmt_invested_row}", _FMT_MULT, kpi=True)

    _auto_width(ws)
    return layout


# ============================================================
# Sheet 6 — Sensitivity (engine-computed snapshots, clearly labelled)
# ============================================================

def _write_sensitivity_grid(ws, r: int, title: str, row_label: str, row_key: str, grid: list[dict], method: str) -> int:
    _style_section_label(ws, r, 1, title)
    r += 1
    axis_values = [c["exit_multiple"] for c in grid[0]["cells"]]
    ws.cell(row=r, column=1, value=f"{row_label} \\ Exit Multiple").font = _FONT_NOTE
    for j, ev in enumerate(axis_values):
        cell = ws.cell(row=r, column=2 + j, value=f"{ev:.2f}x")
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = _ALIGN_CENTER
    r += 1
    for row_entry in grid:
        row_val = row_entry[row_key]
        ws.cell(row=r, column=1, value=f"{row_val:.2f}x").font = _FONT_BOLD
        for j, cell_data in enumerate(row_entry["cells"]):
            irr, moic = cell_data["irr"], cell_data["moic"]
            txt = f"{irr*100:.1f}% / {moic:.2f}x" if irr is not None and moic is not None else "N/A"
            c = ws.cell(row=r, column=2 + j, value=txt)
            c.font = _FONT_NORMAL
            c.alignment = _ALIGN_RIGHT
        r += 1
    note = ws.cell(row=r, column=1, value=method)
    note.font = _FONT_NOTE
    note.alignment = _ALIGN_WRAP
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=1 + len(axis_values))
    ws.row_dimensions[r].height = 30
    r += 2
    return r


def _build_sensitivity_sheet(wb: Workbook, entry_exit: dict | None, exit_leverage: dict | None):
    ws = wb.create_sheet(_SHEET_SENSITIVITY)
    ws.sheet_properties.tabColor = _GOLD
    _write_title(ws, "SENSITIVITY — computed snapshots at export time (see note — not live formulas)", 6)

    r = 3
    note = ws.cell(row=r, column=1, value=(
        "These grids are SNAPSHOTS computed with the same valuation model used everywhere "
        "else in this workbook, at export time. Recomputing a full multi-year debt schedule "
        "per grid cell natively in Excel (25 cells × several years each) is not practical "
        "without VBA/macros, so — unlike every other tab in this workbook — this one is NOT "
        "live: it will not move if you edit the Cover & Assumptions tab. To see a LIVE "
        "recalculation, edit Exit Multiple / Entry Multiple there and watch the Returns & "
        "Waterfall tab instead — that chain IS fully formula-driven."
    ))
    note.font = _FONT_NOTE
    note.alignment = _ALIGN_WRAP
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.row_dimensions[r].height = 60
    r += 3

    if exit_leverage:
        r = _write_sensitivity_grid(
            ws, r, "IRR / MOIC — Leverage × Exit Multiple", "Leverage", "leverage",
            exit_leverage["grid"], exit_leverage["method"],
        )
    if entry_exit:
        r = _write_sensitivity_grid(
            ws, r, "IRR / MOIC — Entry Multiple × Exit Multiple", "Entry Multiple", "entry_multiple",
            entry_exit["grid"], entry_exit["method"],
        )

    _auto_width(ws, min_width=14)


# ============================================================
# Sheet 7 — Credit Metrics (derived from Operating Model / Debt Schedule only)
# ============================================================

def _build_credit_metrics_sheet(
    wb: Workbook, res: dict, alay: AssumptionsLayout, play: PnlLayout, dlay: DebtScheduleLayout, n_years: int,
):
    ws = wb.create_sheet(_SHEET_CREDIT)
    ws.sheet_properties.tabColor = _GOLD

    projs = res.get("projections", [])
    n_cols = n_years + 1
    _write_title(ws, "CREDIT METRICS — derived from Operating Model & Debt Schedule (no new engine data)", n_cols)

    hdr_row = 3
    for idx, p in enumerate(projs):
        ws.cell(row=hdr_row, column=idx + 2, value=f"Year {p['year']}").font = _FONT_HEADER
    _style_header_row(ws, hdr_row, n_cols)

    r = 4
    ws.cell(row=r, column=1, value="Gross Debt / EBITDA").font = _FONT_NORMAL
    for idx in range(len(projs)):
        col = idx + 2
        cell = ws.cell(
            row=r, column=col,
            value=f"={_xref(_SHEET_DEBT, col, dlay.row_total_debt_eoy)}/{_xref(_SHEET_PNL, col, play.row_ebitda)}",
        )
        cell.number_format = _FMT_MULT
        cell.alignment = _ALIGN_RIGHT
        cell.font = _FONT_LINK
    r += 1

    # Tâche "P1" (Partie B) — le cash accumulé par le sweep est désormais
    # suivi (Debt Schedule tab) : le levier NET en tient compte, ce n'est
    # plus uniquement la dette brute.
    ws.cell(row=r, column=1, value="Net Debt / EBITDA").font = _FONT_BOLD
    net_leverage_row = r
    for idx in range(len(projs)):
        col = idx + 2
        cell = ws.cell(
            row=r, column=col,
            value=f"={_xref(_SHEET_DEBT, col, dlay.row_net_debt_eoy)}/{_xref(_SHEET_PNL, col, play.row_ebitda)}",
        )
        cell.number_format = _FMT_MULT
        cell.alignment = _ALIGN_RIGHT
        cell.font = _FONT_LINK_BOLD
    r += 1

    ws.cell(row=r, column=1, value="Leverage Covenant Cap (declining)").font = _FONT_NORMAL
    ws.cell(row=r, column=1).comment = Comment(
        "ESTIMATE — entry net leverage stepped down by a fixed %/year (Cover & Assumptions "
        "tab). Standard mid-market credit-agreement structure, not a re-derivation of any "
        "engine formula. See Checks tab for the PASS/FAIL breach test.", "PE Tracker",
    )
    covenant_cap_row = r
    entry_leverage_ref = _xref(_SHEET_ASSUMPTIONS, 2, alay.row_leverage_mult)
    stepdown_ref = _xref(_SHEET_ASSUMPTIONS, 2, alay.row_leverage_covenant_stepdown)
    for idx in range(len(projs)):
        col = idx + 2
        year = projs[idx]["year"]
        if year == 0:
            cell = ws.cell(row=r, column=col, value="N/A")
        else:
            cell = ws.cell(
                row=r, column=col,
                value=f"=MAX(0.5,{entry_leverage_ref}*(1-{stepdown_ref}*{year}))",
            )
            cell.number_format = _FMT_MULT
        cell.alignment = _ALIGN_RIGHT
        cell.font = _FONT_NORMAL
    r += 1

    ws.cell(row=r, column=1, value="EBITDA / Interest (coverage)").font = _FONT_NORMAL
    for idx in range(len(projs)):
        col = idx + 2
        if idx == 0:
            cell = ws.cell(row=r, column=col, value="N/A")
        else:
            cell = ws.cell(
                row=r, column=col,
                value=(
                    f'=IF({_xref(_SHEET_PNL, col, play.row_interest)}=0,"N/A",'
                    f"{_xref(_SHEET_PNL, col, play.row_ebitda)}/{_xref(_SHEET_PNL, col, play.row_interest)})"
                ),
            )
        cell.number_format = _FMT_MULT
        cell.alignment = _ALIGN_RIGHT
        cell.font = _FONT_LINK
    r += 1

    ws.cell(row=r, column=1, value="FCF / Total Debt").font = _FONT_NORMAL
    for idx in range(len(projs)):
        col = idx + 2
        cell = ws.cell(
            row=r, column=col,
            value=(
                f'=IF({_xref(_SHEET_DEBT, col, dlay.row_total_debt_eoy)}=0,"N/A",'
                f"{_xref(_SHEET_PNL, col, play.row_fcf)}/{_xref(_SHEET_DEBT, col, dlay.row_total_debt_eoy)})"
            ),
        )
        cell.number_format = _FMT_PCT
        cell.alignment = _ALIGN_RIGHT
        cell.font = _FONT_LINK
    r += 1

    ws.cell(row=r, column=1, value="DSCR (CFADS / Debt Service DUE)").font = _FONT_NORMAL
    ws.cell(row=r, column=1).comment = Comment(
        "Tâche \"P1 : physique financière\" (Partie D) — CFADS = EBITDA − Capex − ΔWCR − "
        "Tax (cash available for debt service, BEFORE interest/principal). Debt Service = "
        "Interest + Scheduled Amort. DUE (contractual, from the Debt Schedule tab — NOT the "
        "amount actually paid, which is cash-capped). Before this fix, the denominator used "
        "the amount actually REPAID (cash-capped by construction to equal CFADS), making "
        "this ratio an algebraic identity (always ~1.000x, never able to signal a covenant "
        "breach) — the classic error an IC reviewer catches in 10 seconds. Using the "
        "CONTRACTUAL due amount instead makes this a real ratio that varies year to year and "
        "can fall below the covenant threshold (Checks tab).", "PE Tracker",
    )
    dscr_row = r
    for idx in range(len(projs)):
        col = idx + 2
        if idx == 0:
            cell = ws.cell(row=r, column=col, value="N/A")
        else:
            cfads = (
                f"({_xref(_SHEET_PNL, col, play.row_ebitda)}-{_xref(_SHEET_PNL, col, play.row_capex)}"
                f"-{_xref(_SHEET_PNL, col, play.row_wcr)}-{_xref(_SHEET_PNL, col, play.row_tax)})"
            )
            debt_service_due = (
                f"({_xref(_SHEET_PNL, col, play.row_interest)}+{_xref(_SHEET_DEBT, col, dlay.row_total_sched_due)})"
            )
            cell = ws.cell(row=r, column=col, value=f'=IF({debt_service_due}=0,"N/A",{cfads}/{debt_service_due})')
        cell.number_format = _FMT_MULT
        cell.alignment = _ALIGN_RIGHT
        cell.font = _FONT_LINK_BOLD
    r += 1

    _auto_width(ws, min_width=14)
    return {"net_leverage_row": net_leverage_row, "covenant_cap_row": covenant_cap_row, "dscr_row": dscr_row}


# ============================================================
# Sheet 8 — Checks (audit cells, every PASS/FAIL is a formula)
# ============================================================

def _build_checks_sheet(
    wb: Workbook, res: dict, alay: AssumptionsLayout, play: PnlLayout | None,
    dlay: DebtScheduleLayout | None, returns_layout: dict, credit_layout: dict | None,
    n_years: int, has_debt_schedule: bool, is_v2: bool,
):
    ws = wb.create_sheet(_SHEET_CHECKS)
    ws.sheet_properties.tabColor = _RED
    _write_title(ws, "CHECKS — every PASS/FAIL below is a formula, never a hard-coded value", 4)

    projs = res.get("projections", [])
    n_cols = n_years + 1
    year1_col = 3  # column 2 = Year 0 (excluded from ratio checks: DSCR/leverage are N/A there)

    hdr_row = 3
    for c, h in enumerate(["Check", "Result", "Detail (should be 0 / TRUE)"], start=1):
        cell = ws.cell(row=hdr_row, column=c, value=h)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = _ALIGN_CENTER
    _style_header_row(ws, hdr_row, 3)

    r = hdr_row + 1
    first_check_row = r

    def _pass_fail_cell(row: int, condition_true_formula: str):
        cell = ws.cell(row=row, column=2, value=f'=IF({condition_true_formula},"PASS","FAIL")')
        cell.font = _FONT_BOLD
        cell.alignment = _ALIGN_CENTER
        return cell

    def _rng(sheet: str, row: int) -> str:
        return f"'{sheet}'!{_col(year1_col)}{row}:{_col(n_cols)}{row}"

    # Tâche "P1 : physique financière du modèle LBO" (Partie F) — 3 checks
    # RETIRÉS ici (pas "corrigés") car tautologiques par construction, quels
    # que soient les chiffres réels du modèle — ils n'ont détecté AUCUNE des
    # 6 erreurs du rapport IC : "Sources = Uses" (une equity plug équilibre
    # TOUJOURS son propre total, quoi que Uses contienne — voir Sources &
    # Uses tab pour le rappel visuel, non audité) ; "MOIC × Entry Equity =
    # Exit Equity" (MOIC EST défini comme Exit Equity ÷ Entry Equity — se
    # comparer à sa propre définition ne peut jamais échouer) ; "IRR =
    # MOIC^(1/n)−1" (le closed-form IS la définition même de =IRR() sur une
    # série [-entrée, 0, …, 0, +sortie] — identité algébrique, pas un test).
    # Remplacés par des tests qui PEUVENT échouer et qui, ensemble, auraient
    # détecté les 6 erreurs : suffisance de cash pour le service de la dette
    # contractuel (Erreur 2), DSCR réel sous covenant (Erreur 3), cash
    # minimum jamais négatif (Erreur 1), covenant de levier net décroissant
    # (Partie D).

    # 1. EBITDA consistent between Cover & Assumptions and Operating Model Year 0
    if play is not None:
        ws.cell(row=r, column=1, value="EBITDA consistent (Assumptions vs Operating Model Year 0)").font = _FONT_NORMAL
        diff_formula = f"ROUND({_xref(_SHEET_ASSUMPTIONS, 2, alay.row_entry_ebitda)}-{_xref(_SHEET_PNL, 2, play.row_ebitda)},2)"
        _pass_fail_cell(r, f"{diff_formula}=0")
        detail = ws.cell(row=r, column=3, value=f"={diff_formula}")
        detail.number_format = _FMT_EUR
        r += 1

    # 2. Debt roll-forward per tranche/block: opening[t] == closing[t-1] for all t >= 1
    if has_debt_schedule and dlay is not None:
        tranches_meta = (
            [{"name": "Senior Debt"}] if is_v2 else res.get("debt_tranches_detail", [])
        )
        for t_idx, block in enumerate(dlay.tranche_blocks):
            label = tranches_meta[t_idx].get("name", f"Tranche {t_idx + 1}") if t_idx < len(tranches_meta) else f"Tranche {t_idx + 1}"
            ws.cell(row=r, column=1, value=f"Debt roll-forward consistent — {label}").font = _FONT_NORMAL
            terms = [
                f"({_xref(_SHEET_DEBT, idx + 2, block['opening'])}-{_xref(_SHEET_DEBT, idx + 1, block['closing'])})"
                for idx in range(1, len(projs))
            ]
            diff_formula = f"ROUND({'+'.join(terms)},2)" if terms else "0"
            _pass_fail_cell(r, f"{diff_formula}=0")
            detail = ws.cell(row=r, column=3, value=f"={diff_formula}")
            detail.number_format = _FMT_EUR
            r += 1

        # 3. No negative closing debt balance across any tranche/year
        ws.cell(row=r, column=1, value="No negative debt balance (any tranche, any year)").font = _FONT_NORMAL
        min_parts = [
            "MIN(" + ",".join(f"{_xref(_SHEET_DEBT, idx + 2, b['closing'])}" for idx in range(len(projs))) + ")"
            for b in dlay.tranche_blocks
        ]
        min_formula = f"MIN({','.join(min_parts)})"
        _pass_fail_cell(r, f"ROUND({min_formula},2)>=0")
        detail = ws.cell(row=r, column=3, value=f"={min_formula}")
        detail.number_format = _FMT_EUR
        r += 1

        # 4. Debt service shortfall — CAN fail: any year where cash available
        # did not cover the contractually scheduled amortization (Erreur 2).
        # Before this fix, an unpaid schedule was silently absorbed by a bare
        # MIN() with no check ever looking at the gap between due and paid.
        ws.cell(row=r, column=1, value="No debt-service shortfall (any year — DUE fully PAID)").font = _FONT_NORMAL
        shortfall_range = _rng(_SHEET_DEBT, dlay.row_total_shortfall)
        _pass_fail_cell(r, f"ROUND(MAX({shortfall_range}),2)=0")
        detail = ws.cell(row=r, column=3, value=f"=MAX({shortfall_range})")
        detail.number_format = _FMT_EUR
        r += 1

        # 5. Minimum cash never negative (any year) — a real test: it can
        # only ever fail if the Assumptions inputs (fees/min cash %) are set
        # to something nonsensical, but it is a genuine, non-tautological
        # test of the cash-balance mechanic introduced in Partie B.
        ws.cell(row=r, column=1, value="Cash balance never negative (any year)").font = _FONT_NORMAL
        cash_range = f"'{_SHEET_DEBT}'!{_col(2)}{dlay.row_cash_balance_eoy}:{_col(n_cols)}{dlay.row_cash_balance_eoy}"
        _pass_fail_cell(r, f"ROUND(MIN({cash_range}),2)>=0")
        detail = ws.cell(row=r, column=3, value=f"=MIN({cash_range})")
        detail.number_format = _FMT_EUR
        r += 1

    # 6. Exit Equity = Exit EV - Net Debt, cross-checked against the Debt
    # Schedule tab's OWN independently-derived net debt (not the Returns
    # sheet's own copies of the same two cells restated — that would be the
    # same tautology as before under a new name). This can genuinely fail if
    # the two tabs' debt/cash chains ever diverge.
    rl = returns_layout
    ws.cell(row=r, column=1, value="Exit Equity = Exit EV − Net Debt (cross-checked vs Debt Schedule tab)").font = _FONT_NORMAL
    if has_debt_schedule and dlay is not None:
        last_col_letter = _col(n_cols)
        independent_net_debt = f"'{_SHEET_DEBT}'!{last_col_letter}{dlay.row_net_debt_eoy}"
    else:
        independent_net_debt = _xref(_SHEET_RETURNS, 2, rl["exit_net_debt_row"])
    diff_formula = (
        f"ROUND({_xref(_SHEET_RETURNS, 2, rl['exit_eq_row'])}-"
        f"({_xref(_SHEET_RETURNS, 2, rl['exit_ev_row'])}-{independent_net_debt}),2)"
    )
    _pass_fail_cell(r, f"{diff_formula}=0")
    detail = ws.cell(row=r, column=3, value=f"={diff_formula}")
    detail.number_format = _FMT_EUR
    r += 1

    # 7. DSCR ≥ covenant minimum, every year — CAN fail: a real ratio (Partie
    # D fix), not the identity it replaces. Falls below threshold whenever
    # CFADS does not comfortably cover interest + contractual amortization.
    if credit_layout is not None:
        ws.cell(row=r, column=1, value="DSCR ≥ covenant minimum (every year)").font = _FONT_NORMAL
        dscr_range = _rng(_SHEET_CREDIT, credit_layout["dscr_row"])
        covenant_ref = _xref(_SHEET_ASSUMPTIONS, 2, alay.row_dscr_covenant_min)
        _pass_fail_cell(r, f"SUMPRODUCT(--({dscr_range}<{covenant_ref}))=0")
        detail = ws.cell(row=r, column=3, value=f"=MIN({dscr_range})")
        detail.number_format = _FMT_MULT
        r += 1

        # 8. Leverage covenant — net leverage under the declining cap, every
        # year. CAN fail: exactly the mechanism that flags a distressed case
        # (see verification: a stress scenario trips this to FAIL).
        ws.cell(row=r, column=1, value="Leverage covenant respected (net debt/EBITDA ≤ declining cap)").font = _FONT_NORMAL
        net_lev_range = _rng(_SHEET_CREDIT, credit_layout["net_leverage_row"])
        cap_range = _rng(_SHEET_CREDIT, credit_layout["covenant_cap_row"])
        _pass_fail_cell(r, f"SUMPRODUCT(--({net_lev_range}>{cap_range}))=0")
        detail = ws.cell(row=r, column=3, value=f"=MAX({net_lev_range}-{cap_range})")
        detail.number_format = _FMT_MULT
        r += 1

    last_check_row = r - 1

    r += 1
    ws.cell(row=r, column=1, value="OVERALL").font = _FONT_SECTION
    ws.cell(row=r, column=1).fill = _FILL_LIGHT
    overall_cell = ws.cell(
        row=r, column=2,
        value=f'=IF(COUNTIF(B{first_check_row}:B{last_check_row},"FAIL")=0,"ALL CHECKS PASSED","REVIEW REQUIRED")',
    )
    overall_cell.font = _FONT_SECTION
    overall_cell.fill = _FILL_LIGHT
    overall_cell.alignment = _ALIGN_CENTER

    # Conditional formatting colours the PASS/FAIL text green/red based on
    # the cell's OWN computed value — never a static colour applied
    # regardless of outcome (that would silently lie about a FAIL).
    result_range = f"B{first_check_row}:B{last_check_row}"
    ws.conditional_formatting.add(
        result_range,
        CellIsRule(operator="equal", formula=['"PASS"'], font=Font(color=_GREEN, bold=True)),
    )
    ws.conditional_formatting.add(
        result_range,
        CellIsRule(operator="equal", formula=['"FAIL"'], font=Font(color=_RED, bold=True)),
    )
    ws.conditional_formatting.add(
        f"B{r}",
        CellIsRule(operator="equal", formula=['"ALL CHECKS PASSED"'], font=Font(color=_GREEN, bold=True)),
    )
    ws.conditional_formatting.add(
        f"B{r}",
        CellIsRule(operator="equal", formula=['"REVIEW REQUIRED"'], font=Font(color=_RED, bold=True)),
    )

    # _auto_width FIRST — it unconditionally overwrites every column's width
    # from content length, so an explicit override must come AFTER it (found
    # during verification: column A's check labels were being silently
    # shrunk back down by the auto-width pass that used to run last).
    _auto_width(ws, min_width=14)
    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["C"].width = 18
